"""Excel brigade piecework report generator"""
from typing import Optional, List, Dict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from .excel_print_form_generator import ExcelPrintFormGenerator
from ..data.database_manager import DatabaseManager


class ExcelBrigadePieceworkReport(ExcelPrintFormGenerator):
    """Generator for brigade piecework report in Excel format"""
    
    TEMPLATE_NAME = "brigade_piecework_template.xlsx"
    
    def __init__(self):
        """Initialize Excel brigade piecework report generator"""
        super().__init__()
        self.db = DatabaseManager().get_connection()
    
    def generate(self, period_start: str, period_end: str, filters: dict = None) -> Optional[bytes]:
        """
        Generate brigade piecework report in Excel format
        
        Args:
            period_start: Start date (YYYY-MM-DD)
            period_end: End date (YYYY-MM-DD)
            filters: Optional filters (object_id, estimate_id, work_id, executor_id)
            
        Returns:
            Excel content as bytes
        """
        if filters is None:
            filters = {}
        
        # Load report data
        report_data = self._load_report_data(period_start, period_end, filters)
        if not report_data:
            return None
        
        # Create workbook
        workbook = self._create_report(report_data, period_start, period_end)
        
        return self.save_to_bytes(workbook)
    
    def _load_report_data(self, period_start: str, period_end: str, filters: dict) -> Optional[dict]:
        """Load report data from database"""
        cursor = self.db.cursor()
        
        # Build WHERE clause
        where_parts = ["wer.period BETWEEN ? AND ?"]
        params = [period_start, period_end]
        
        if filters.get('object_id'):
            where_parts.append("e.object_id = ?")
            params.append(filters['object_id'])
        
        if filters.get('estimate_id'):
            where_parts.append("wer.estimate_id = ?")
            params.append(filters['estimate_id'])
        
        if filters.get('work_id'):
            where_parts.append("wer.work_id = ?")
            params.append(filters['work_id'])
        
        where_clause = " AND ".join(where_parts)
        
        # Handle executor filter separately for fact queries
        executor_filter = ""
        executor_params = []
        if filters.get('executor_id'):
            executor_filter = " AND wer.recorder_type = 'daily_report' AND wer.recorder_id IN (SELECT id FROM daily_reports WHERE foreman_id = ?)"
            executor_params = [filters['executor_id']]
        
        # Get header info
        cursor.execute(f"""
            SELECT 
                o.name as object_name,
                e.number as estimate_number,
                e.date as estimate_date
            FROM work_execution_register wer
            LEFT JOIN estimates e ON wer.estimate_id = e.id
            LEFT JOIN objects o ON e.object_id = o.id
            WHERE {where_clause}
            LIMIT 1
        """, params)
        
        header_row = cursor.fetchone()
        if not header_row:
            return None
        
        # Get executor name if filtered
        executor_name = "Все исполнители"
        if filters.get('executor_id'):
            cursor.execute("SELECT full_name FROM persons WHERE id = ?", (filters['executor_id'],))
            executor_row = cursor.fetchone()
            if executor_row:
                executor_name = executor_row['full_name']
        
        # Get all unique periods (dates)
        cursor.execute(f"""
            SELECT DISTINCT wer.period
            FROM work_execution_register wer
            WHERE {where_clause}
            ORDER BY wer.period
        """, params)
        
        periods = [row['period'] for row in cursor.fetchall()]
        
        # Get plan quantities per work
        plan_query = f"""
            SELECT 
                w.id as work_id,
                w.name as work_name,
                w.unit,
                COALESCE(
                    (SELECT el.price FROM estimate_lines el 
                     WHERE el.estimate_id = wer.estimate_id AND el.work_id = wer.work_id 
                     LIMIT 1), 
                    0
                ) as price,
                SUM(wer.quantity_income) as plan_quantity
            FROM work_execution_register wer
            LEFT JOIN works w ON wer.work_id = w.id
            WHERE {where_clause} AND wer.quantity_income > 0
            GROUP BY w.id
            ORDER BY w.name
        """
        
        cursor.execute(plan_query, params)
        plan_rows = cursor.fetchall()
        
        # Initialize works data with plan
        works_data = {}
        for row in plan_rows:
            work_id = row['work_id']
            works_data[work_id] = {
                'work_name': row['work_name'],
                'executor_name': executor_name,
                'unit': row['unit'] or '',
                'price': row['price'] or 0,
                'plan_quantity': row['plan_quantity'] or 0,
                'fact_quantity': 0,
                'periods': {}
            }
        
        # Get fact quantities by period
        fact_query = f"""
            SELECT 
                w.id as work_id,
                wer.period,
                SUM(wer.quantity_expense) as fact_quantity,
                wer.recorder_type,
                wer.recorder_id
            FROM work_execution_register wer
            LEFT JOIN works w ON wer.work_id = w.id
            WHERE {where_clause} AND wer.quantity_expense > 0 {executor_filter}
            GROUP BY w.id, wer.period
            ORDER BY w.name, wer.period
        """
        
        cursor.execute(fact_query, params + executor_params)
        fact_rows = cursor.fetchall()
        
        # Get executor names from daily reports
        executor_names_by_work = {}
        if fact_rows:
            for row in fact_rows:
                if row['recorder_type'] == 'daily_report':
                    cursor.execute("""
                        SELECT p.full_name 
                        FROM daily_reports dr
                        JOIN persons p ON dr.foreman_id = p.id
                        WHERE dr.id = ?
                    """, (row['recorder_id'],))
                    executor_row = cursor.fetchone()
                    if executor_row and row['work_id'] not in executor_names_by_work:
                        executor_names_by_work[row['work_id']] = executor_row['full_name']
        
        # Fill fact data
        for row in fact_rows:
            work_id = row['work_id']
            if work_id not in works_data:
                continue
            
            # Update executor name if available
            if work_id in executor_names_by_work:
                works_data[work_id]['executor_name'] = executor_names_by_work[work_id]
            
            # Accumulate fact by period
            period = row['period']
            if period not in works_data[work_id]['periods']:
                works_data[work_id]['periods'][period] = {
                    'quantity': 0,
                    'sum': 0
                }
            
            fact_qty = row['fact_quantity'] or 0
            works_data[work_id]['periods'][period]['quantity'] += fact_qty
            works_data[work_id]['periods'][period]['sum'] += fact_qty * works_data[work_id]['price']
            works_data[work_id]['fact_quantity'] += fact_qty
        
        return {
            'object_name': header_row['object_name'] or '',
            'estimate_number': header_row['estimate_number'] or '',
            'estimate_date': header_row['estimate_date'] or '',
            'executor_name': executor_name,
            'period_start': period_start,
            'period_end': period_end,
            'periods': periods,
            'works': list(works_data.values())
        }
    
    def _create_report(self, data: dict, period_start: str, period_end: str) -> Workbook:
        """Create Excel report from scratch"""
        workbook = self.create_workbook()
        sheet = workbook.active
        sheet.title = "Отчёт по проекту"
        
        # Styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        current_row = 1
        
        # Title
        title = f'Выполнение работ по проекту "{data["object_name"]}" {data["executor_name"]} - сдельная'
        self.set_cell_value(sheet, current_row, 1, title)
        sheet.merge_cells(f'A{current_row}:E{current_row}')
        self.set_cell_style(sheet, current_row, 1, 
                          font=Font(bold=True, size=14),
                          alignment=Alignment(horizontal='left', vertical='center'))
        current_row += 1
        
        # Period
        period_text = f'c {self.format_date(period_start)} по {self.format_date(period_end)}'
        self.set_cell_value(sheet, current_row, 1, period_text)
        sheet.merge_cells(f'A{current_row}:E{current_row}')
        self.set_cell_style(sheet, current_row, 1,
                          font=Font(size=11),
                          alignment=Alignment(horizontal='left', vertical='center'))
        current_row += 1
        
        # Headers
        header_row1 = current_row
        header_row2 = current_row + 1
        
        # Fixed columns headers
        fixed_headers = [
            ('A', '№ п/п'),
            ('B', 'Наименование работ'),
            ('C', 'Исполнитель'),
            ('D', 'Ед. изм.'),
            ('E', 'Стоимость на ед.')
        ]
        
        for col_letter, header_text in fixed_headers:
            sheet.merge_cells(f'{col_letter}{header_row1}:{col_letter}{header_row2}')
            cell = sheet[f'{col_letter}{header_row1}']
            cell.value = header_text
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = header_fill
            cell.border = thin_border
        
        # Plan/Fact summary columns
        col_idx = 6
        summary_start_col = col_idx
        
        # Plan columns
        for offset, header in enumerate(['План', 'Факт', '+/-', '%']):
            col_letter = get_column_letter(col_idx + offset)
            cell = sheet[f'{col_letter}{header_row2}']
            cell.value = header
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = header_fill
            cell.border = thin_border
        
        sheet.merge_cells(f'{get_column_letter(col_idx)}{header_row1}:{get_column_letter(col_idx+3)}{header_row1}')
        cell = sheet[f'{get_column_letter(col_idx)}{header_row1}']
        cell.value = 'Объём по проекту'
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = header_fill
        cell.border = thin_border
        
        col_idx += 4
        
        # Cost summary columns
        for offset, header in enumerate(['План', 'Факт', '+/-', '%']):
            col_letter = get_column_letter(col_idx + offset)
            cell = sheet[f'{col_letter}{header_row2}']
            cell.value = header
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = header_fill
            cell.border = thin_border
        
        sheet.merge_cells(f'{get_column_letter(col_idx)}{header_row1}:{get_column_letter(col_idx+3)}{header_row1}')
        cell = sheet[f'{get_column_letter(col_idx)}{header_row1}']
        cell.value = 'Стоимость по проекту'
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = header_fill
        cell.border = thin_border
        
        col_idx += 4
        
        # Period columns (dates)
        periods = data['periods']
        for period in periods:
            # Date header
            period_date = datetime.strptime(period, '%Y-%m-%d')
            col_letter = get_column_letter(col_idx)
            sheet.merge_cells(f'{col_letter}{header_row1}:{get_column_letter(col_idx+1)}{header_row1}')
            cell = sheet[f'{col_letter}{header_row1}']
            cell.value = period_date
            cell.number_format = 'DD.MM.YYYY'
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = header_fill
            cell.border = thin_border
            
            # Subheaders
            for offset, subheader in enumerate(['Факт объёмы', 'Факт cтоимость']):
                col_letter = get_column_letter(col_idx + offset)
                cell = sheet[f'{col_letter}{header_row2}']
                cell.value = subheader
                cell.font = Font(bold=True, size=8)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.fill = header_fill
                cell.border = thin_border
            
            col_idx += 2
        
        current_row = header_row2 + 1
        
        # Set column widths
        sheet.column_dimensions['A'].width = 6
        sheet.column_dimensions['B'].width = 35
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 10
        sheet.column_dimensions['E'].width = 12
        
        for i in range(summary_start_col, col_idx):
            sheet.column_dimensions[get_column_letter(i)].width = 10
        
        # Data rows
        total_plan_qty = 0
        total_fact_qty = 0
        total_plan_sum = 0
        total_fact_sum = 0
        total_by_period = {period: {'quantity': 0, 'sum': 0} for period in periods}
        
        for idx, work in enumerate(data['works'], 1):
            row_num = current_row
            
            # Fixed columns
            self.set_cell_value(sheet, row_num, 1, idx)
            self.set_cell_value(sheet, row_num, 2, work['work_name'])
            self.set_cell_value(sheet, row_num, 3, work['executor_name'])
            self.set_cell_value(sheet, row_num, 4, work['unit'])
            self.set_cell_value(sheet, row_num, 5, work['price'])
            
            # Plan/Fact quantities
            plan_qty = work['plan_quantity']
            fact_qty = work['fact_quantity']
            diff_qty = fact_qty - plan_qty
            pct_qty = (fact_qty / plan_qty * 100) if plan_qty > 0 else 0
            
            self.set_cell_value(sheet, row_num, 6, plan_qty)
            self.set_cell_value(sheet, row_num, 7, fact_qty)
            self.set_cell_value(sheet, row_num, 8, diff_qty)
            self.set_cell_value(sheet, row_num, 9, pct_qty)
            
            # Plan/Fact sums
            plan_sum = plan_qty * work['price']
            fact_sum = fact_qty * work['price']
            diff_sum = fact_sum - plan_sum
            pct_sum = (fact_sum / plan_sum * 100) if plan_sum > 0 else 0
            
            self.set_cell_value(sheet, row_num, 10, plan_sum)
            self.set_cell_value(sheet, row_num, 11, fact_sum)
            self.set_cell_value(sheet, row_num, 12, diff_sum)
            self.set_cell_value(sheet, row_num, 13, pct_sum)
            
            # Accumulate totals
            total_plan_qty += plan_qty
            total_fact_qty += fact_qty
            total_plan_sum += plan_sum
            total_fact_sum += fact_sum
            
            # Period data
            col_idx = 14
            for period in periods:
                period_data = work['periods'].get(period, {'quantity': 0, 'sum': 0})
                self.set_cell_value(sheet, row_num, col_idx, period_data['quantity'])
                self.set_cell_value(sheet, row_num, col_idx + 1, period_data['sum'])
                
                # Accumulate period totals
                total_by_period[period]['quantity'] += period_data['quantity']
                total_by_period[period]['sum'] += period_data['sum']
                
                col_idx += 2
            
            # Apply borders
            for col in range(1, col_idx):
                cell = sheet.cell(row_num, col)
                cell.border = thin_border
                if col >= 5:
                    cell.number_format = '#,##0.00'
                if col == 9 or col == 13:
                    cell.number_format = '0.00"%"'
            
            current_row += 1
        
        # Total row
        total_row = current_row
        self.set_cell_value(sheet, total_row, 1, 'Итого')
        sheet.merge_cells(f'A{total_row}:D{total_row}')
        
        # Total quantities
        total_diff_qty = total_fact_qty - total_plan_qty
        total_pct_qty = (total_fact_qty / total_plan_qty * 100) if total_plan_qty > 0 else 0
        
        self.set_cell_value(sheet, total_row, 6, total_plan_qty)
        self.set_cell_value(sheet, total_row, 7, total_fact_qty)
        self.set_cell_value(sheet, total_row, 8, total_diff_qty)
        self.set_cell_value(sheet, total_row, 9, total_pct_qty)
        
        # Total sums
        total_diff_sum = total_fact_sum - total_plan_sum
        total_pct_sum = (total_fact_sum / total_plan_sum * 100) if total_plan_sum > 0 else 0
        
        self.set_cell_value(sheet, total_row, 10, total_plan_sum)
        self.set_cell_value(sheet, total_row, 11, total_fact_sum)
        self.set_cell_value(sheet, total_row, 12, total_diff_sum)
        self.set_cell_value(sheet, total_row, 13, total_pct_sum)
        
        # Period totals
        col_idx = 14
        for period in periods:
            self.set_cell_value(sheet, total_row, col_idx, total_by_period[period]['quantity'])
            self.set_cell_value(sheet, total_row, col_idx + 1, total_by_period[period]['sum'])
            col_idx += 2
        
        # Apply total row styling
        for col in range(1, col_idx):
            cell = sheet.cell(total_row, col)
            cell.font = Font(bold=True, size=11)
            cell.fill = total_fill
            cell.border = thin_border
            if col >= 5:
                cell.number_format = '#,##0.00'
            if col == 9 or col == 13:
                cell.number_format = '0.00"%"'
        
        return workbook
