"""Excel estimate print form generator with resource statement"""
from typing import Optional, List, Dict
from openpyxl.styles import Font, Alignment, Border, Side
from .excel_print_form_generator import ExcelPrintFormGenerator
from ..data.database_manager import DatabaseManager


class ExcelEstimateResourcePrintForm(ExcelPrintFormGenerator):
    """Generator for estimate print forms with resource statement in Excel format"""
    
    TEMPLATE_NAME = "estimate_resource_template.xlsx"
    
    def __init__(self):
        """Initialize Excel estimate resource print form generator"""
        super().__init__()
        self.db_manager = DatabaseManager()
        self.db_manager.initialize('construction.db')
    
    def generate(self, estimate_id: int) -> Optional[bytes]:
        """
        Generate estimate print form with resource statement in Excel format
        
        Args:
            estimate_id: ID of the estimate
            
        Returns:
            Excel content as bytes or None if estimate not found
        """
        # Load estimate data
        estimate_data = self._load_estimate_data(estimate_id)
        if not estimate_data:
            return None
        
        # Load resource data
        resource_data = self._load_resource_data(estimate_id)
        
        # Try to load template, otherwise create from scratch
        if self.template_exists(self.TEMPLATE_NAME):
            workbook = self.load_template(self.TEMPLATE_NAME)
            if workbook:
                self._fill_template(workbook, estimate_data, resource_data)
            else:
                workbook = self._create_from_scratch(estimate_data, resource_data)
        else:
            workbook = self._create_from_scratch(estimate_data, resource_data)
        
        return self.save_to_bytes(workbook)
    
    def _load_estimate_data(self, estimate_id: int) -> Optional[dict]:
        """Load estimate data from database"""
        try:
            with self.db_manager.get_session() as session:
                from src.data.models.sqlalchemy_models import (
                    Estimate, EstimateLine, Work, Counterparty, Object, Organization, Person
                )
                
                # Load estimate header
                estimate = session.query(Estimate).filter(Estimate.id == estimate_id).first()
                if not estimate:
                    return None
                
                estimate_data = {
                    'id': estimate.id,
                    'number': estimate.number,
                    'date': estimate.date,
                    'total_sum': estimate.total_sum or 0,
                    'total_labor': estimate.total_labor or 0,
                    'customer_name': estimate.customer.name if estimate.customer else None,
                    'object_name': estimate.object.name if estimate.object else None,
                    'contractor_name': estimate.contractor.name if estimate.contractor else None,
                    'responsible_name': estimate.responsible.full_name if estimate.responsible else None,
                    'lines': []
                }
                
                # Load estimate lines
                lines = session.query(EstimateLine).filter(
                    EstimateLine.estimate_id == estimate_id
                ).order_by(EstimateLine.line_number).all()
                
                for line in lines:
                    # Check if this is a group row (work_id = -1)
                    is_group = line.work_id == -1
                    
                    estimate_data['lines'].append({
                        'line_number': line.line_number,
                        'work_name': line.work.name if line.work and not is_group else (line.unit if not is_group else ""),
                        'work_code': line.work.code if line.work else "",
                        'quantity': line.quantity or 0,
                        'unit': line.unit if not is_group else "",
                        'price': line.price or 0,
                        'labor_rate': line.labor_rate or 0,
                        'sum': line.sum or 0,
                        'planned_labor': line.planned_labor or 0,
                        'is_group': is_group,
                        'group_name': line.unit if is_group else ""
                    })
                
                return estimate_data
                
        except Exception as e:
            print(f"Error loading estimate data: {e}")
            return None
    
    def _load_resource_data(self, estimate_id: int) -> List[Dict]:
        """Load resource data (materials) for the estimate"""
        try:
            with self.db_manager.get_session() as session:
                from src.data.models.sqlalchemy_models import (
                    EstimateLine, Work, CostItemMaterial, Material
                )
                from sqlalchemy import func
                
                # Get materials from works used in estimate
                query = session.query(
                    Material.id,
                    Material.code,
                    Material.description,
                    Material.price,
                    Material.unit,
                    func.sum(CostItemMaterial.quantity_per_unit * EstimateLine.quantity).label('total_quantity'),
                    func.sum(CostItemMaterial.quantity_per_unit * EstimateLine.quantity * Material.price).label('total_sum')
                ).select_from(EstimateLine)\
                .join(Work, EstimateLine.work_id == Work.id)\
                .join(CostItemMaterial, CostItemMaterial.work_id == Work.id)\
                .join(Material, CostItemMaterial.material_id == Material.id)\
                .filter(
                    EstimateLine.estimate_id == estimate_id,
                    EstimateLine.work_id != -1,  # Exclude group rows
                    CostItemMaterial.material_id.isnot(None)
                ).group_by(
                    Material.id, Material.code, Material.description, Material.price, Material.unit
                ).order_by(Material.code)
                
                resources = []
                for idx, row in enumerate(query.all(), 1):
                    resources.append({
                        'line_number': idx,
                        'code': row.code or '',
                        'description': row.description or '',
                        'unit': row.unit or '',
                        'quantity': float(row.total_quantity or 0),
                        'price': float(row.price or 0),
                        'sum': float(row.total_sum or 0)
                    })
                
                return resources
                
        except Exception as e:
            print(f"Error loading resource data: {e}")
            return []
    
    def _fill_template(self, workbook, estimate_data: dict, resource_data: List[Dict]):
        """Fill template with data"""
        # Get or create sheets
        if len(workbook.worksheets) >= 2:
            estimate_sheet = workbook.worksheets[0]
            resource_sheet = workbook.worksheets[1]
        else:
            estimate_sheet = workbook.active
            resource_sheet = workbook.create_sheet("Ресурсная ведомость")
        
        # Fill estimate sheet
        self._fill_estimate_sheet(estimate_sheet, estimate_data)
        
        # Fill resource sheet
        self._fill_resource_sheet(resource_sheet, estimate_data, resource_data)
    
    def _fill_estimate_sheet(self, sheet, estimate_data: dict):
        """Fill estimate sheet with data"""
        # Try to find and fill placeholders
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    value = str(cell.value)
                    # Replace placeholders
                    value = value.replace('{NUMBER}', estimate_data['number'] or '')
                    value = value.replace('{DATE}', self.format_date(estimate_data['date']))
                    value = value.replace('{OBJECT}', estimate_data['object_name'] or '')
                    value = value.replace('{CUSTOMER}', estimate_data['customer_name'] or '')
                    value = value.replace('{CONTRACTOR}', estimate_data['contractor_name'] or '')
                    value = value.replace('{TOTAL_SUM}', self.format_number(estimate_data['total_sum']))
                    value = value.replace('{TOTAL_LABOR}', self.format_number(estimate_data['total_labor']))
                    cell.value = value
        
        # Find table start (look for header row with "N" or "№")
        table_start_row = None
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=50), start=1):
            for cell in row:
                if cell.value and str(cell.value).strip() in ['N', '№', 'N п/п', '№ п/п']:
                    table_start_row = row_idx + 1
                    break
            if table_start_row:
                break
        
        # Fill lines if table found
        if table_start_row:
            self._fill_lines_in_template(sheet, estimate_data['lines'], table_start_row)
    
    def _fill_resource_sheet(self, sheet, estimate_data: dict, resource_data: List[Dict]):
        """Fill resource sheet with data"""
        sheet.title = "Ресурсная ведомость"
        
        # Clear existing content
        sheet.delete_rows(1, sheet.max_row)
        
        current_row = 1
        
        # Title
        self.set_cell_value(sheet, current_row, 1, "РЕСУРСНАЯ ВЕДОМОСТЬ")
        self.merge_cells(sheet, current_row, 1, current_row, 5)
        title_style = self.create_title_style()
        self.set_cell_style(sheet, current_row, 1, **title_style)
        current_row += 1
        
        self.set_cell_value(sheet, current_row, 1, f"к смете №{estimate_data['number']}")
        self.merge_cells(sheet, current_row, 1, current_row, 5)
        self.set_cell_style(sheet, current_row, 1,
                           font=Font(size=10, italic=True),
                           alignment=Alignment(horizontal='center', vertical='center'))
        current_row += 2
        
        # Table header
        headers = ["№ п/п", "Материал", "Количество", "Цена", "Сумма"]
        for col, header in enumerate(headers, start=1):
            self.set_cell_value(sheet, current_row, col, header)
            header_style = self.create_header_style()
            self.set_cell_style(sheet, current_row, col, **header_style)
        current_row += 1
        
        # Set column widths
        widths = [8, 60, 15, 12, 15]
        for col, width in enumerate(widths, start=1):
            self.set_column_width(sheet, col, width)
        
        # Data rows
        data_style = self.create_data_style()
        number_style = self.create_number_style()
        
        total_sum = 0.0
        for resource in resource_data:
            # Format material description with code
            material_desc = f"{resource['code']} - {resource['description']}"
            quantity_text = f"{self.format_number(resource['quantity'], 3)} {resource['unit']}"
            
            self.set_cell_value(sheet, current_row, 1, resource['line_number'])
            self.set_cell_value(sheet, current_row, 2, material_desc)
            self.set_cell_value(sheet, current_row, 3, quantity_text)
            self.set_cell_value(sheet, current_row, 4, self.safe_float(resource['price']))
            self.set_cell_value(sheet, current_row, 5, self.safe_float(resource['sum']))
            
            # Apply styles
            self.set_cell_style(sheet, current_row, 1, **data_style)
            self.set_cell_style(sheet, current_row, 2, **data_style)
            self.set_cell_style(sheet, current_row, 3, **data_style)
            self.set_cell_style(sheet, current_row, 4, **number_style)
            self.set_cell_style(sheet, current_row, 5, **number_style)
            
            total_sum += resource['sum']
            current_row += 1
        
        # Total row
        current_row += 1
        self.set_cell_value(sheet, current_row, 4, "ИТОГО:")
        self.set_cell_value(sheet, current_row, 5, self.safe_float(total_sum))
        
        # Style total row
        total_style = {
            'font': Font(bold=True, size=11),
            'alignment': Alignment(horizontal='right', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        self.set_cell_style(sheet, current_row, 4, **total_style)
        self.set_cell_style(sheet, current_row, 5, **total_style)
    
    def _fill_lines_in_template(self, sheet, lines: list, start_row: int):
        """Fill estimate lines in template"""
        current_row = start_row
        
        for line in lines:
            if line.get('is_group', False):
                # Group row
                self.set_cell_value(sheet, current_row, 1, line['line_number'])
                self.set_cell_value(sheet, current_row, 2, line.get('group_name', ''))
                # Make group row bold
                for col in range(1, 14):
                    cell = sheet.cell(row=current_row, column=col)
                    cell.font = Font(bold=True, size=10)
            else:
                # Regular line
                unit_labor = line['labor_rate'] if line['quantity'] > 0 else 0
                unit_materials = (line['price'] - unit_labor) if line['price'] > unit_labor else 0
                unit_total = line['price']
                unit_tz = line['labor_rate'] / line['quantity'] if line['quantity'] > 0 else 0
                
                total_labor = unit_labor * line['quantity']
                total_materials = unit_materials * line['quantity']
                total_sum = line['sum']
                total_tz = line['planned_labor']
                
                self.set_cell_value(sheet, current_row, 1, line['line_number'])
                self.set_cell_value(sheet, current_row, 2, line.get('work_code', ''))
                self.set_cell_value(sheet, current_row, 3, line['work_name'])
                self.set_cell_value(sheet, current_row, 4, line['unit'])
                self.set_cell_value(sheet, current_row, 5, self.safe_float(line['quantity']))
                self.set_cell_value(sheet, current_row, 6, self.safe_float(unit_labor))
                self.set_cell_value(sheet, current_row, 7, self.safe_float(unit_materials))
                self.set_cell_value(sheet, current_row, 8, self.safe_float(unit_total))
                self.set_cell_value(sheet, current_row, 9, self.safe_float(unit_tz))
                self.set_cell_value(sheet, current_row, 10, self.safe_float(total_labor))
                self.set_cell_value(sheet, current_row, 11, self.safe_float(total_materials))
                self.set_cell_value(sheet, current_row, 12, self.safe_float(total_sum))
                self.set_cell_value(sheet, current_row, 13, self.safe_float(total_tz))
            
            current_row += 1
    
    def _create_from_scratch(self, estimate_data: dict, resource_data: List[Dict]):
        """Create Excel document from scratch with two sheets"""
        workbook = self.create_workbook()
        
        # Create estimate sheet
        estimate_sheet = workbook.active
        estimate_sheet.title = "Смета"
        self._create_estimate_sheet(estimate_sheet, estimate_data)
        
        # Create resource sheet
        resource_sheet = workbook.create_sheet("Ресурсная ведомость")
        self._create_resource_sheet(resource_sheet, estimate_data, resource_data)
        
        return workbook
    
    def _create_estimate_sheet(self, sheet, estimate_data: dict):
        """Create estimate sheet from scratch"""
        current_row = 1
        
        # Approval section
        self.set_cell_value(sheet, current_row, 1, "Согласовано:")
        self.set_cell_value(sheet, current_row, 6, "Утверждаю:")
        current_row += 1
        
        customer = estimate_data['customer_name'] or "_______________"
        contractor = estimate_data['contractor_name'] or "_______________"
        self.set_cell_value(sheet, current_row, 1, f"Заказчик: {customer}")
        self.set_cell_value(sheet, current_row, 6, f"Подрядчик: {contractor}")
        current_row += 3
        
        # Title
        title_row = current_row
        self.set_cell_value(sheet, title_row, 1, f"ЛОКАЛЬНЫЙ СМЕТНЫЙ РАСЧЕТ №{estimate_data['number']}")
        self.merge_cells(sheet, title_row, 1, title_row, 13)
        title_style = self.create_title_style()
        self.set_cell_style(sheet, title_row, 1, **title_style)
        current_row += 1
        
        self.set_cell_value(sheet, current_row, 1, "(локальная смета)")
        self.merge_cells(sheet, current_row, 1, current_row, 13)
        self.set_cell_style(sheet, current_row, 1, 
                           alignment=Alignment(horizontal='center', vertical='center'))
        current_row += 2
        
        # Object
        object_desc = estimate_data['object_name'] or "Объект не указан"
        self.set_cell_value(sheet, current_row, 1, object_desc)
        self.merge_cells(sheet, current_row, 1, current_row, 13)
        self.set_cell_style(sheet, current_row, 1,
                           alignment=Alignment(horizontal='center', vertical='center'))
        current_row += 1
        
        self.set_cell_value(sheet, current_row, 1, "(наименование работ и затрат, наименование объекта)")
        self.merge_cells(sheet, current_row, 1, current_row, 13)
        self.set_cell_style(sheet, current_row, 1,
                           font=Font(size=9, italic=True),
                           alignment=Alignment(horizontal='center', vertical='center'))
        current_row += 2
        
        # Table header
        header_row1 = current_row
        headers1 = ["", "", "", "", "", "на единицу работ", "", "", "", "на полный объём работ", "", "", ""]
        for col, header in enumerate(headers1, start=1):
            self.set_cell_value(sheet, header_row1, col, header)
            header_style = self.create_header_style()
            self.set_cell_style(sheet, header_row1, col, **header_style)
        
        self.merge_cells(sheet, header_row1, 6, header_row1, 9)
        self.merge_cells(sheet, header_row1, 10, header_row1, 13)
        current_row += 1
        
        header_row2 = current_row
        headers2 = ["N", "Код", "Наименование работ и затрат", "ЕдИзм", "Кол-во",
                   "Зарплата", "Материалы,\nмеханизмы", "Всего", "ТЗ",
                   "Зарплата", "Материалы,\nмеханизмы", "Всего", "ТЗ"]
        for col, header in enumerate(headers2, start=1):
            self.set_cell_value(sheet, header_row2, col, header)
            header_style = self.create_header_style()
            self.set_cell_style(sheet, header_row2, col, **header_style)
        current_row += 1
        
        # Set column widths
        widths = [5, 10, 35, 8, 8, 10, 10, 10, 8, 10, 10, 10, 8]
        for col, width in enumerate(widths, start=1):
            self.set_column_width(sheet, col, width)
        
        # Data rows
        data_style = self.create_data_style()
        number_style = self.create_number_style()
        
        for line in estimate_data['lines']:
            if line.get('is_group', False):
                # Group row
                self.set_cell_value(sheet, current_row, 1, line['line_number'])
                self.set_cell_value(sheet, current_row, 2, line.get('group_name', ''))
                self.merge_cells(sheet, current_row, 2, current_row, 13)
                for col in range(1, 14):
                    cell = sheet.cell(row=current_row, column=col)
                    cell.font = Font(bold=True, size=10)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            else:
                # Regular line
                unit_labor = line['labor_rate'] if line['quantity'] > 0 else 0
                unit_materials = (line['price'] - unit_labor) if line['price'] > unit_labor else 0
                unit_total = line['price']
                unit_tz = line['labor_rate'] / line['quantity'] if line['quantity'] > 0 else 0
                
                total_labor = unit_labor * line['quantity']
                total_materials = unit_materials * line['quantity']
                total_sum = line['sum']
                total_tz = line['planned_labor']
                
                self.set_cell_value(sheet, current_row, 1, line['line_number'])
                self.set_cell_value(sheet, current_row, 2, line.get('work_code', ''))
                self.set_cell_value(sheet, current_row, 3, line['work_name'])
                self.set_cell_value(sheet, current_row, 4, line['unit'])
                self.set_cell_value(sheet, current_row, 5, self.safe_float(line['quantity']))
                self.set_cell_value(sheet, current_row, 6, self.safe_float(unit_labor))
                self.set_cell_value(sheet, current_row, 7, self.safe_float(unit_materials))
                self.set_cell_value(sheet, current_row, 8, self.safe_float(unit_total))
                self.set_cell_value(sheet, current_row, 9, self.safe_float(unit_tz))
                self.set_cell_value(sheet, current_row, 10, self.safe_float(total_labor))
                self.set_cell_value(sheet, current_row, 11, self.safe_float(total_materials))
                self.set_cell_value(sheet, current_row, 12, self.safe_float(total_sum))
                self.set_cell_value(sheet, current_row, 13, self.safe_float(total_tz))
                
                # Apply styles
                for col in range(1, 14):
                    if col in [1, 2, 3, 4]:
                        self.set_cell_style(sheet, current_row, col, **data_style)
                    else:
                        self.set_cell_style(sheet, current_row, col, **number_style)
            
            current_row += 1
        
        # Totals
        current_row += 1
        self.set_cell_value(sheet, current_row, 1, "ИТОГО по смете:")
        self.set_cell_style(sheet, current_row, 1, font=Font(bold=True, size=11))
        current_row += 1
        
        self.set_cell_value(sheet, current_row, 1, 
                           f"Всего: {self.format_number(estimate_data['total_sum'])} руб.")
        current_row += 1
        
        self.set_cell_value(sheet, current_row, 1,
                           f"Трудозатраты: {self.format_number(estimate_data['total_labor'])} ч.")
    
    def _create_resource_sheet(self, sheet, estimate_data: dict, resource_data: List[Dict]):
        """Create resource sheet from scratch"""
        current_row = 1
        
        # Title
        self.set_cell_value(sheet, current_row, 1, "РЕСУРСНАЯ ВЕДОМОСТЬ")
        self.merge_cells(sheet, current_row, 1, current_row, 5)
        title_style = self.create_title_style()
        self.set_cell_style(sheet, current_row, 1, **title_style)
        current_row += 1
        
        self.set_cell_value(sheet, current_row, 1, f"к смете №{estimate_data['number']}")
        self.merge_cells(sheet, current_row, 1, current_row, 5)
        self.set_cell_style(sheet, current_row, 1,
                           font=Font(size=10, italic=True),
                           alignment=Alignment(horizontal='center', vertical='center'))
        current_row += 2
        
        # Table header
        headers = ["№ п/п", "Материал", "Количество", "Цена", "Сумма"]
        for col, header in enumerate(headers, start=1):
            self.set_cell_value(sheet, current_row, col, header)
            header_style = self.create_header_style()
            self.set_cell_style(sheet, current_row, col, **header_style)
        current_row += 1
        
        # Set column widths
        widths = [8, 60, 15, 12, 15]
        for col, width in enumerate(widths, start=1):
            self.set_column_width(sheet, col, width)
        
        # Data rows
        data_style = self.create_data_style()
        number_style = self.create_number_style()
        
        total_sum = 0.0
        for resource in resource_data:
            # Format material description with code
            material_desc = f"{resource['code']} - {resource['description']}"
            quantity_text = f"{self.format_number(resource['quantity'], 3)} {resource['unit']}"
            
            self.set_cell_value(sheet, current_row, 1, resource['line_number'])
            self.set_cell_value(sheet, current_row, 2, material_desc)
            self.set_cell_value(sheet, current_row, 3, quantity_text)
            self.set_cell_value(sheet, current_row, 4, self.safe_float(resource['price']))
            self.set_cell_value(sheet, current_row, 5, self.safe_float(resource['sum']))
            
            # Apply styles
            self.set_cell_style(sheet, current_row, 1, **data_style)
            self.set_cell_style(sheet, current_row, 2, **data_style)
            self.set_cell_style(sheet, current_row, 3, **data_style)
            self.set_cell_style(sheet, current_row, 4, **number_style)
            self.set_cell_style(sheet, current_row, 5, **number_style)
            
            total_sum += resource['sum']
            current_row += 1
        
        # Total row
        current_row += 1
        self.set_cell_value(sheet, current_row, 4, "ИТОГО:")
        self.set_cell_value(sheet, current_row, 5, self.safe_float(total_sum))
        
        # Style total row
        total_style = {
            'font': Font(bold=True, size=11),
            'alignment': Alignment(horizontal='right', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        self.set_cell_style(sheet, current_row, 4, **total_style)
        self.set_cell_style(sheet, current_row, 5, **total_style)
    
    def create_template(self) -> bool:
        """Create default template file with two sheets"""
        try:
            # Create a sample template with placeholders
            workbook = self.create_workbook()
            
            # Create estimate sheet
            estimate_sheet = workbook.active
            estimate_sheet.title = "Смета"
            
            # Add placeholders and structure for estimate
            self.set_cell_value(estimate_sheet, 1, 1, "Согласовано:")
            self.set_cell_value(estimate_sheet, 1, 6, "Утверждаю:")
            self.set_cell_value(estimate_sheet, 2, 1, "Заказчик: {CUSTOMER}")
            self.set_cell_value(estimate_sheet, 2, 6, "Подрядчик: {CONTRACTOR}")
            
            self.set_cell_value(estimate_sheet, 5, 1, "ЛОКАЛЬНЫЙ СМЕТНЫЙ РАСЧЕТ №{NUMBER}")
            self.merge_cells(estimate_sheet, 5, 1, 5, 13)
            title_style = self.create_title_style()
            self.set_cell_style(estimate_sheet, 5, 1, **title_style)
            
            self.set_cell_value(estimate_sheet, 6, 1, "(локальная смета)")
            self.merge_cells(estimate_sheet, 6, 1, 6, 13)
            
            self.set_cell_value(estimate_sheet, 8, 1, "{OBJECT}")
            self.merge_cells(estimate_sheet, 8, 1, 8, 13)
            
            # Table header for estimate
            header_row1 = 11
            headers1 = ["", "", "", "", "", "на единицу работ", "", "", "", "на полный объём работ", "", "", ""]
            for col, header in enumerate(headers1, start=1):
                self.set_cell_value(estimate_sheet, header_row1, col, header)
                header_style = self.create_header_style()
                self.set_cell_style(estimate_sheet, header_row1, col, **header_style)
            
            self.merge_cells(estimate_sheet, header_row1, 6, header_row1, 9)
            self.merge_cells(estimate_sheet, header_row1, 10, header_row1, 13)
            
            header_row2 = 12
            headers2 = ["N", "Код", "Наименование работ и затрат", "ЕдИзм", "Кол-во",
                       "Зарплата", "Материалы,\nмеханизмы", "Всего", "ТЗ",
                       "Зарплата", "Материалы,\nмеханизмы", "Всего", "ТЗ"]
            for col, header in enumerate(headers2, start=1):
                self.set_cell_value(estimate_sheet, header_row2, col, header)
                header_style = self.create_header_style()
                self.set_cell_style(estimate_sheet, header_row2, col, **header_style)
            
            # Set column widths for estimate
            widths = [5, 10, 35, 8, 8, 10, 10, 10, 8, 10, 10, 10, 8]
            for col, width in enumerate(widths, start=1):
                self.set_column_width(estimate_sheet, col, width)
            
            # Create resource sheet
            resource_sheet = workbook.create_sheet("Ресурсная ведомость")
            
            # Add structure for resource sheet
            self.set_cell_value(resource_sheet, 1, 1, "РЕСУРСНАЯ ВЕДОМОСТЬ")
            self.merge_cells(resource_sheet, 1, 1, 1, 5)
            self.set_cell_style(resource_sheet, 1, 1, **title_style)
            
            self.set_cell_value(resource_sheet, 2, 1, "к смете №{NUMBER}")
            self.merge_cells(resource_sheet, 2, 1, 2, 5)
            
            # Resource table header
            resource_headers = ["№ п/п", "Материал", "Количество", "Цена", "Сумма"]
            for col, header in enumerate(resource_headers, start=1):
                self.set_cell_value(resource_sheet, 4, col, header)
                header_style = self.create_header_style()
                self.set_cell_style(resource_sheet, 4, col, **header_style)
            
            # Set column widths for resource sheet
            resource_widths = [8, 60, 15, 12, 15]
            for col, width in enumerate(resource_widths, start=1):
                self.set_column_width(resource_sheet, col, width)
            
            # Save template
            template_path = self.get_template_path(self.TEMPLATE_NAME)
            workbook.save(template_path)
            return True
        except Exception as e:
            print(f"Error creating template: {e}")
            return False