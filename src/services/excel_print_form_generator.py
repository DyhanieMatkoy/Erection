"""Excel print form generator using templates"""
import os
import configparser
from typing import Optional, Dict, Any
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO


class ExcelPrintFormGenerator:
    """Base class for generating print forms in Excel format"""
    
    def __init__(self):
        """Initialize Excel print form generator"""
        self.templates_path = self._get_templates_path()
        self._ensure_templates_directory()
    
    def _get_templates_path(self) -> str:
        """Get templates path from config"""
        config = configparser.ConfigParser()
        if os.path.exists('env.ini'):
            config.read('env.ini', encoding='utf-8')
            if 'PrintForms' in config and 'templates_path' in config['PrintForms']:
                return config['PrintForms']['templates_path']
        return 'PrnForms'
    
    def _ensure_templates_directory(self):
        """Ensure templates directory exists"""
        if not os.path.exists(self.templates_path):
            os.makedirs(self.templates_path)
    
    def get_template_path(self, template_name: str) -> str:
        """Get full path to template file"""
        return os.path.join(self.templates_path, template_name)
    
    def template_exists(self, template_name: str) -> bool:
        """Check if template file exists"""
        return os.path.exists(self.get_template_path(template_name))
    
    def load_template(self, template_name: str) -> Optional[Workbook]:
        """Load Excel template"""
        template_path = self.get_template_path(template_name)
        if not os.path.exists(template_path):
            return None
        try:
            return load_workbook(template_path)
        except Exception as e:
            print(f"Error loading template {template_name}: {e}")
            return None
    
    def create_workbook(self) -> Workbook:
        """Create new workbook"""
        return Workbook()
    
    def save_to_bytes(self, workbook: Workbook) -> bytes:
        """Save workbook to bytes"""
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def format_date(self, date_obj) -> str:
        """Format date object to string"""
        if date_obj:
            if isinstance(date_obj, str):
                try:
                    parsed = datetime.strptime(date_obj, "%Y-%m-%d")
                    return parsed.strftime("%d.%m.%Y")
                except:
                    return date_obj
            return date_obj.strftime("%d.%m.%Y")
        return ""
    
    def format_number(self, value: float, decimals: int = 2) -> str:
        """Format number with specified decimal places"""
        if value is None:
            return "0." + "0" * decimals
        return f"{value:.{decimals}f}"
    
    def safe_float(self, value) -> float:
        """Safely convert value to float, returning 0.0 for None values"""
        if value is None:
            return 0.0
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0
    
    def set_cell_value(self, sheet, row: int, col: int, value: Any):
        """Set cell value"""
        sheet.cell(row=row, column=col, value=value)
    
    def set_cell_style(self, sheet, row: int, col: int, 
                       font: Font = None, alignment: Alignment = None,
                       border: Border = None, fill: PatternFill = None):
        """Set cell style"""
        cell = sheet.cell(row=row, column=col)
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
        if fill:
            cell.fill = fill
    
    def create_header_style(self) -> Dict[str, Any]:
        """Create header cell style"""
        return {
            'font': Font(bold=True, size=11),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'fill': PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        }
    
    def create_data_style(self) -> Dict[str, Any]:
        """Create data cell style"""
        return {
            'font': Font(size=10),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    def create_number_style(self) -> Dict[str, Any]:
        """Create number cell style"""
        style = self.create_data_style()
        style['alignment'] = Alignment(horizontal='right', vertical='center')
        return style
    
    def create_title_style(self) -> Dict[str, Any]:
        """Create title cell style"""
        return {
            'font': Font(bold=True, size=14),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
        }
    
    def merge_cells(self, sheet, start_row: int, start_col: int, end_row: int, end_col: int):
        """Merge cells"""
        start_cell = f"{get_column_letter(start_col)}{start_row}"
        end_cell = f"{get_column_letter(end_col)}{end_row}"
        sheet.merge_cells(f"{start_cell}:{end_cell}")
    
    def set_column_width(self, sheet, col: int, width: float):
        """Set column width"""
        sheet.column_dimensions[get_column_letter(col)].width = width
    
    def set_row_height(self, sheet, row: int, height: float):
        """Set row height"""
        sheet.row_dimensions[row].height = height
