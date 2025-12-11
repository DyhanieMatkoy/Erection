import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from typing import List, Dict, Optional, Tuple
from decimal import Decimal

class WorkSpecificationExcelService:
    """Service for importing and exporting work specifications via Excel"""
    
    def export_specifications(self, specifications: List[Dict], file_path: str) -> bool:
        """
        Export specifications to Excel file
        
        Args:
            specifications: List of specification dictionaries
            file_path: Output file path
            
        Returns:
            True if successful, False otherwise
        """
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Спецификация"
            
            # Headers
            headers = ["Тип", "Наименование", "Ед.изм.", "Норма расхода", "Цена", "Стоимость"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                
            # Data
            for row_idx, spec in enumerate(specifications, 2):
                ws.cell(row=row_idx, column=1, value=spec.get('component_type', ''))
                ws.cell(row=row_idx, column=2, value=spec.get('component_name', ''))
                
                # Unit (name if available, else ID)
                # Ideally we should resolve unit name before passing here, or pass it in spec dict
                # The WorkForm sets unit_id. The repo returns unit_id.
                # If unit_name is not in spec, we might export ID or empty.
                # For now let's assume spec might have 'unit_name' injected or we just export ID/empty.
                # Requirement 7.2 doesn't specify unit resolution but implies user readable.
                # We'll use whatever is in 'unit_name' or 'unit_id'.
                unit_val = spec.get('unit_name', str(spec.get('unit_id', '')))
                ws.cell(row=row_idx, column=3, value=unit_val)
                
                rate = float(spec.get('consumption_rate', 0))
                ws.cell(row=row_idx, column=4, value=rate)
                
                price = float(spec.get('unit_price', 0))
                ws.cell(row=row_idx, column=5, value=price)
                
                total = float(spec.get('total_cost', rate * price))
                ws.cell(row=row_idx, column=6, value=total)
                
            # Column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            
            wb.save(file_path)
            return True
        except Exception as e:
            print(f"Export error: {e}")
            return False
            
    def import_specifications(self, file_path: str) -> Tuple[List[Dict], str]:
        """
        Import specifications from Excel file
        
        Args:
            file_path: Input file path
            
        Returns:
            Tuple of (List of specification dictionaries, error message)
        """
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            specs = []
            
            # Identify headers
            # Assume first row is header
            # Map columns by name
            header_map = {}
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=col).value
                if val:
                    val = str(val).lower().strip()
                    if "тип" in val:
                        header_map['type'] = col
                    elif "наименование" in val:
                        header_map['name'] = col
                    elif "ед" in val: # ед.изм., единица
                        header_map['unit'] = col
                    elif "норма" in val or "расход" in val:
                        header_map['rate'] = col
                    elif "цена" in val:
                        header_map['price'] = col
            
            if 'name' not in header_map:
                return [], "Не найден столбец 'Наименование'"
                
            for row in range(2, ws.max_row + 1):
                name_cell = ws.cell(row=row, column=header_map['name'])
                if not name_cell.value:
                    continue
                    
                spec = {}
                spec['component_name'] = str(name_cell.value).strip()
                
                if 'type' in header_map:
                    spec['component_type'] = str(ws.cell(row=row, column=header_map['type']).value or 'Material').strip()
                else:
                    spec['component_type'] = 'Material' # Default
                    
                # Validate type
                if spec['component_type'] not in ['Material', 'Labor', 'Equipment', 'Other']:
                    spec['component_type'] = 'Material'
                    
                if 'unit' in header_map:
                    spec['unit_name'] = str(ws.cell(row=row, column=header_map['unit']).value or '').strip()
                    # We need to resolve unit ID later
                    
                if 'rate' in header_map:
                    try:
                        val = ws.cell(row=row, column=header_map['rate']).value
                        spec['consumption_rate'] = float(val) if val is not None else 0.0
                    except:
                        spec['consumption_rate'] = 0.0
                else:
                    spec['consumption_rate'] = 0.0
                    
                if 'price' in header_map:
                    try:
                        val = ws.cell(row=row, column=header_map['price']).value
                        spec['unit_price'] = float(val) if val is not None else 0.0
                    except:
                        spec['unit_price'] = 0.0
                else:
                    spec['unit_price'] = 0.0
                    
                specs.append(spec)
                
            return specs, ""
        except Exception as e:
            return [], str(e)
