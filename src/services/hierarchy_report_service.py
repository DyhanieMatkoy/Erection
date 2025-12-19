"""Hierarchy reporting service"""
from typing import Optional, Dict, Any, List
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from ..data.repositories.estimate_hierarchy_repository import EstimateHierarchyRepository
from ..data.repositories.estimate_repository import EstimateRepository
from ..data.models.estimate import EstimateType, HierarchyTree

logger = logging.getLogger(__name__)


class HierarchyReportService:
    def __init__(self):
        self.hierarchy_repo = EstimateHierarchyRepository()
        self.estimate_repo = EstimateRepository()

    def get_hierarchy_tree(self, root_id: int) -> Optional[HierarchyTree]:
        """Get hierarchy tree for a given root estimate"""
        return self.hierarchy_repo.get_hierarchy_tree(root_id)

    def generate_excel_report(self, root_id: int) -> Optional[bytes]:
        """
        Generate Excel report for estimate hierarchy.
        
        Args:
            root_id: ID of the general estimate (root)
            
        Returns:
            Excel content as bytes or None if not found
        """
        tree = self.get_hierarchy_tree(root_id)
        if not tree:
            logger.warning(f"Hierarchy tree not found for root {root_id}")
            return None
            
        wb = Workbook()
        ws = wb.active
        ws.title = "Структура сметы"
        
        # Styles
        header_font = Font(bold=True, size=12)
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        fill_general = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        fill_plan = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        # Headers
        headers = ["Уровень", "Тип", "Номер", "Дата", "Наименование/Объект", "Сумма (руб)", "Трудозатраты (ч)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            
        # Column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        
        current_row = 2
        
        # Process Root (General Estimate)
        root = tree.root
        self._write_node_row(ws, current_row, root, 0, border, fill_general)
        current_row += 1
        
        # Process Children (Plan Estimates)
        if root.children:
            for child in root.children:
                self._write_node_row(ws, current_row, child, 1, border, fill_plan)
                current_row += 1
                
        # Add Summary Section
        current_row += 2
        ws.cell(row=current_row, column=1, value="Сводка").font = header_font
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="Всего смет:").font = bold_font
        ws.cell(row=current_row, column=2, value=tree.total_nodes)
        current_row += 1
        
        total_plan_sum = sum(c.estimate.total_sum for c in root.children)
        total_plan_labor = sum(c.estimate.total_labor for c in root.children)
        
        ws.cell(row=current_row, column=1, value="Сумма плановых смет:").font = bold_font
        ws.cell(row=current_row, column=2, value=total_plan_sum).number_format = '#,##0.00'
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="Трудозатраты плановых:").font = bold_font
        ws.cell(row=current_row, column=2, value=total_plan_labor).number_format = '#,##0.00'
        
        # Check for discrepancies (if needed)
        diff_sum = root.estimate.total_sum - total_plan_sum
        if abs(diff_sum) > 0.01:
            current_row += 1
            ws.cell(row=current_row, column=1, value="Отклонение (Ген - План):").font = Font(bold=True, color="FF0000")
            ws.cell(row=current_row, column=2, value=diff_sum).font = Font(color="FF0000")
            ws.cell(row=current_row, column=2).number_format = '#,##0.00'

        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _write_node_row(self, ws, row_idx, node, level, border, fill):
        estimate = node.estimate
        
        # Indentation visual
        indent = "    " * level
        name_display = f"{indent}{estimate.object_name or 'Без названия'}"
        
        type_name = "Генеральная" if estimate.estimate_type == EstimateType.GENERAL.value else "Плановая"
        
        # Write cells
        cells = [
            (1, level),
            (2, type_name),
            (3, estimate.number),
            (4, estimate.date.strftime("%d.%m.%Y") if estimate.date else ""),
            (5, name_display),
            (6, estimate.total_sum),
            (7, estimate.total_labor)
        ]
        
        for col, value in cells:
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.fill = fill
            
            if col in [6, 7]: # Numbers
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
