"""
Table part import service for handling Excel and CSV imports.

This service provides functionality to import data into table parts
from Excel (.xlsx, .xls) and CSV files with column mapping and validation.
"""

import openpyxl
import csv
import os
from typing import Dict, List, Optional, Tuple, Any, Union
from dataclasses import dataclass
from enum import Enum
from datetime import datetime, date


class ImportFormat(Enum):
    """Supported import file formats"""
    EXCEL = "excel"
    CSV = "csv"


class ImportValidationError(Exception):
    """Exception raised when import validation fails"""
    pass


@dataclass
class ImportColumn:
    """Configuration for an import column"""
    source_name: str  # Column name in import file
    target_field: str  # Target field name in table part
    data_type: str  # Expected data type (str, int, float, date, bool)
    required: bool = False
    default_value: Any = None
    validation_rules: Optional[Dict[str, Any]] = None


@dataclass
class ImportPreview:
    """Preview data for import confirmation"""
    headers: List[str]
    sample_rows: List[List[Any]]
    total_rows: int
    detected_format: ImportFormat
    suggested_mappings: Dict[str, str]  # source_column -> target_field


@dataclass
class ImportResult:
    """Result of import operation"""
    success: bool
    imported_rows: int
    failed_rows: int
    errors: List[str]
    warnings: List[str]
    data: List[Dict[str, Any]]


class TablePartImportService:
    """Service for importing data into table parts"""
    
    def __init__(self):
        self.supported_formats = [ImportFormat.EXCEL, ImportFormat.CSV]
        self.max_preview_rows = 10
        self.max_import_rows = 10000
    
    def detect_file_format(self, file_path: str) -> ImportFormat:
        """Detect file format based on extension"""
        ext = os.path.splitext(file_path)[1].lower()
        if ext in ['.xlsx', '.xls']:
            return ImportFormat.EXCEL
        elif ext == '.csv':
            return ImportFormat.CSV
        else:
            raise ImportValidationError(f"Неподдерживаемый формат файла: {ext}")
    
    def create_preview(self, file_path: str, target_columns: List[ImportColumn]) -> ImportPreview:
        """
        Create preview of import data with suggested column mappings.
        
        Args:
            file_path: Path to the import file
            target_columns: List of target columns for mapping
            
        Returns:
            ImportPreview with headers, sample data, and suggested mappings
        """
        format_type = self.detect_file_format(file_path)
        
        if format_type == ImportFormat.EXCEL:
            return self._create_excel_preview(file_path, target_columns)
        elif format_type == ImportFormat.CSV:
            return self._create_csv_preview(file_path, target_columns)
        else:
            raise ImportValidationError(f"Неподдерживаемый формат: {format_type}")
    
    def import_data(
        self, 
        file_path: str, 
        column_mappings: Dict[str, str],
        target_columns: List[ImportColumn],
        skip_header: bool = True
    ) -> ImportResult:
        """
        Import data from file using specified column mappings.
        
        Args:
            file_path: Path to the import file
            column_mappings: Mapping from source columns to target fields
            target_columns: Target column definitions
            skip_header: Whether to skip the first row as header
            
        Returns:
            ImportResult with success status and imported data
        """
        format_type = self.detect_file_format(file_path)
        
        if format_type == ImportFormat.EXCEL:
            return self._import_excel_data(file_path, column_mappings, target_columns, skip_header)
        elif format_type == ImportFormat.CSV:
            return self._import_csv_data(file_path, column_mappings, target_columns, skip_header)
        else:
            raise ImportValidationError(f"Неподдерживаемый формат: {format_type}")
    
    def _create_excel_preview(self, file_path: str, target_columns: List[ImportColumn]) -> ImportPreview:
        """Create preview for Excel file"""
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value is not None:
                    headers.append(str(cell_value).strip())
                else:
                    headers.append(f"Колонка {col}")
            
            # Get sample rows
            sample_rows = []
            for row in range(2, min(self.max_preview_rows + 2, ws.max_row + 1)):
                row_data = []
                for col in range(1, len(headers) + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    row_data.append(cell_value)
                sample_rows.append(row_data)
            
            # Calculate total rows (excluding header)
            total_rows = max(0, ws.max_row - 1)
            
            # Suggest column mappings
            suggested_mappings = self._suggest_column_mappings(headers, target_columns)
            
            return ImportPreview(
                headers=headers,
                sample_rows=sample_rows,
                total_rows=total_rows,
                detected_format=ImportFormat.EXCEL,
                suggested_mappings=suggested_mappings
            )
            
        except Exception as e:
            raise ImportValidationError(f"Ошибка чтения Excel файла: {str(e)}")
    
    def _create_csv_preview(self, file_path: str, target_columns: List[ImportColumn]) -> ImportPreview:
        """Create preview for CSV file"""
        try:
            # Try different encodings
            encodings = ['utf-8', 'cp1251', 'windows-1251', 'iso-8859-1']
            
            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding, newline='') as file:
                        # Detect delimiter
                        sample = file.read(1024)
                        file.seek(0)
                        
                        sniffer = csv.Sniffer()
                        delimiter = sniffer.sniff(sample).delimiter
                        
                        reader = csv.reader(file, delimiter=delimiter)
                        
                        # Get headers
                        headers = next(reader)
                        headers = [str(h).strip() for h in headers]
                        
                        # Get sample rows
                        sample_rows = []
                        total_rows = 0
                        
                        for i, row in enumerate(reader):
                            total_rows += 1
                            if i < self.max_preview_rows:
                                sample_rows.append(row)
                        
                        # Suggest column mappings
                        suggested_mappings = self._suggest_column_mappings(headers, target_columns)
                        
                        return ImportPreview(
                            headers=headers,
                            sample_rows=sample_rows,
                            total_rows=total_rows,
                            detected_format=ImportFormat.CSV,
                            suggested_mappings=suggested_mappings
                        )
                        
                except UnicodeDecodeError:
                    continue
            
            raise ImportValidationError("Не удалось определить кодировку CSV файла")
            
        except Exception as e:
            raise ImportValidationError(f"Ошибка чтения CSV файла: {str(e)}")
    
    def _suggest_column_mappings(self, headers: List[str], target_columns: List[ImportColumn]) -> Dict[str, str]:
        """Suggest column mappings based on header names"""
        mappings = {}
        
        # Create mapping rules for common column names
        mapping_rules = {
            # Russian column names
            'наименование': ['name', 'title', 'description'],
            'название': ['name', 'title'],
            'код': ['code', 'id'],
            'шифр': ['code', 'cipher'],
            'количество': ['quantity', 'amount'],
            'кол-во': ['quantity', 'amount'],
            'объем': ['volume', 'quantity'],
            'цена': ['price', 'cost'],
            'стоимость': ['cost', 'price'],
            'сумма': ['sum', 'total', 'amount'],
            'единица': ['unit', 'measure'],
            'ед.изм': ['unit', 'measure'],
            'примечание': ['note', 'comment', 'remark'],
            'дата': ['date'],
            'номер': ['number', 'num'],
            
            # English column names
            'name': ['name', 'title'],
            'code': ['code', 'id'],
            'quantity': ['quantity', 'amount'],
            'price': ['price', 'cost'],
            'cost': ['cost', 'price'],
            'sum': ['sum', 'total'],
            'total': ['total', 'sum'],
            'unit': ['unit', 'measure'],
            'date': ['date'],
            'number': ['number', 'num'],
            'note': ['note', 'comment'],
        }
        
        # Try to match headers with target columns
        for header in headers:
            header_lower = header.lower().strip()
            
            # Direct match first
            for target_col in target_columns:
                if header_lower == target_col.target_field.lower():
                    mappings[header] = target_col.target_field
                    break
            
            # If no direct match, try fuzzy matching
            if header not in mappings:
                for rule_key, rule_targets in mapping_rules.items():
                    if rule_key in header_lower:
                        for target_col in target_columns:
                            if target_col.target_field.lower() in rule_targets:
                                mappings[header] = target_col.target_field
                                break
                        if header in mappings:
                            break
        
        return mappings
    
    def _import_excel_data(
        self, 
        file_path: str, 
        column_mappings: Dict[str, str],
        target_columns: List[ImportColumn],
        skip_header: bool
    ) -> ImportResult:
        """Import data from Excel file"""
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Get headers
            headers = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value is not None:
                    headers.append(str(cell_value).strip())
                else:
                    headers.append(f"Колонка {col}")
            
            # Create target column lookup
            target_col_lookup = {col.target_field: col for col in target_columns}
            
            # Import data
            imported_data = []
            errors = []
            warnings = []
            imported_rows = 0
            failed_rows = 0
            
            start_row = 2 if skip_header else 1
            
            for row_idx in range(start_row, ws.max_row + 1):
                try:
                    row_data = {}
                    has_data = False
                    
                    # Extract data for mapped columns
                    for col_idx, header in enumerate(headers):
                        if header in column_mappings:
                            target_field = column_mappings[header]
                            cell_value = ws.cell(row=row_idx, column=col_idx + 1).value
                            
                            if cell_value is not None:
                                has_data = True
                                
                                # Convert and validate value
                                if target_field in target_col_lookup:
                                    target_col = target_col_lookup[target_field]
                                    converted_value = self._convert_value(
                                        cell_value, target_col.data_type, target_field
                                    )
                                    row_data[target_field] = converted_value
                                else:
                                    row_data[target_field] = cell_value
                    
                    # Skip empty rows
                    if not has_data:
                        continue
                    
                    # Validate required fields
                    validation_errors = self._validate_row_data(row_data, target_columns)
                    if validation_errors:
                        failed_rows += 1
                        errors.extend([f"Строка {row_idx}: {err}" for err in validation_errors])
                        continue
                    
                    # Add default values for missing fields
                    for target_col in target_columns:
                        if target_col.target_field not in row_data and target_col.default_value is not None:
                            row_data[target_col.target_field] = target_col.default_value
                    
                    imported_data.append(row_data)
                    imported_rows += 1
                    
                    # Check import limit
                    if imported_rows >= self.max_import_rows:
                        warnings.append(f"Импорт ограничен {self.max_import_rows} строками")
                        break
                        
                except Exception as e:
                    failed_rows += 1
                    errors.append(f"Строка {row_idx}: {str(e)}")
            
            return ImportResult(
                success=imported_rows > 0,
                imported_rows=imported_rows,
                failed_rows=failed_rows,
                errors=errors,
                warnings=warnings,
                data=imported_data
            )
            
        except Exception as e:
            return ImportResult(
                success=False,
                imported_rows=0,
                failed_rows=0,
                errors=[f"Ошибка импорта: {str(e)}"],
                warnings=[],
                data=[]
            )
    
    def _import_csv_data(
        self, 
        file_path: str, 
        column_mappings: Dict[str, str],
        target_columns: List[ImportColumn],
        skip_header: bool
    ) -> ImportResult:
        """Import data from CSV file"""
        try:
            # Try different encodings
            encodings = ['utf-8', 'cp1251', 'windows-1251', 'iso-8859-1']
            
            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding, newline='') as file:
                        # Detect delimiter
                        sample = file.read(1024)
                        file.seek(0)
                        
                        sniffer = csv.Sniffer()
                        delimiter = sniffer.sniff(sample).delimiter
                        
                        reader = csv.reader(file, delimiter=delimiter)
                        
                        # Get headers
                        headers = next(reader) if skip_header else []
                        if not skip_header:
                            file.seek(0)
                            reader = csv.reader(file, delimiter=delimiter)
                        
                        # Create target column lookup
                        target_col_lookup = {col.target_field: col for col in target_columns}
                        
                        # Import data
                        imported_data = []
                        errors = []
                        warnings = []
                        imported_rows = 0
                        failed_rows = 0
                        
                        for row_idx, row in enumerate(reader, start=1):
                            try:
                                row_data = {}
                                has_data = False
                                
                                # Extract data for mapped columns
                                if skip_header and headers:
                                    for col_idx, header in enumerate(headers):
                                        if col_idx < len(row) and header in column_mappings:
                                            target_field = column_mappings[header]
                                            cell_value = row[col_idx].strip() if row[col_idx] else None
                                            
                                            if cell_value:
                                                has_data = True
                                                
                                                # Convert and validate value
                                                if target_field in target_col_lookup:
                                                    target_col = target_col_lookup[target_field]
                                                    converted_value = self._convert_value(
                                                        cell_value, target_col.data_type, target_field
                                                    )
                                                    row_data[target_field] = converted_value
                                                else:
                                                    row_data[target_field] = cell_value
                                else:
                                    # No headers, use column mappings by index
                                    for col_idx, cell_value in enumerate(row):
                                        col_name = f"Колонка {col_idx + 1}"
                                        if col_name in column_mappings:
                                            target_field = column_mappings[col_name]
                                            cell_value = cell_value.strip() if cell_value else None
                                            
                                            if cell_value:
                                                has_data = True
                                                
                                                if target_field in target_col_lookup:
                                                    target_col = target_col_lookup[target_field]
                                                    converted_value = self._convert_value(
                                                        cell_value, target_col.data_type, target_field
                                                    )
                                                    row_data[target_field] = converted_value
                                                else:
                                                    row_data[target_field] = cell_value
                                
                                # Skip empty rows
                                if not has_data:
                                    continue
                                
                                # Validate required fields
                                validation_errors = self._validate_row_data(row_data, target_columns)
                                if validation_errors:
                                    failed_rows += 1
                                    errors.extend([f"Строка {row_idx}: {err}" for err in validation_errors])
                                    continue
                                
                                # Add default values for missing fields
                                for target_col in target_columns:
                                    if target_col.target_field not in row_data and target_col.default_value is not None:
                                        row_data[target_col.target_field] = target_col.default_value
                                
                                imported_data.append(row_data)
                                imported_rows += 1
                                
                                # Check import limit
                                if imported_rows >= self.max_import_rows:
                                    warnings.append(f"Импорт ограничен {self.max_import_rows} строками")
                                    break
                                    
                            except Exception as e:
                                failed_rows += 1
                                errors.append(f"Строка {row_idx}: {str(e)}")
                        
                        return ImportResult(
                            success=imported_rows > 0,
                            imported_rows=imported_rows,
                            failed_rows=failed_rows,
                            errors=errors,
                            warnings=warnings,
                            data=imported_data
                        )
                        
                except UnicodeDecodeError:
                    continue
            
            raise ImportValidationError("Не удалось определить кодировку CSV файла")
            
        except Exception as e:
            return ImportResult(
                success=False,
                imported_rows=0,
                failed_rows=0,
                errors=[f"Ошибка импорта: {str(e)}"],
                warnings=[],
                data=[]
            )
    
    def _convert_value(self, value: Any, data_type: str, field_name: str) -> Any:
        """Convert value to specified data type"""
        if value is None or value == '':
            return None
        
        try:
            if data_type == 'str':
                return str(value).strip()
            elif data_type == 'int':
                if isinstance(value, (int, float)):
                    return int(value)
                return int(float(str(value).replace(',', '.')))
            elif data_type == 'float':
                if isinstance(value, (int, float)):
                    return float(value)
                return float(str(value).replace(',', '.'))
            elif data_type == 'bool':
                if isinstance(value, bool):
                    return value
                str_val = str(value).lower().strip()
                return str_val in ['true', '1', 'да', 'yes', 'истина']
            elif data_type == 'date':
                if isinstance(value, date):
                    return value
                elif isinstance(value, datetime):
                    return value.date()
                else:
                    # Try to parse date string
                    import re
                    date_str = str(value).strip()
                    
                    # Try different date formats
                    formats = [
                        r'(\d{1,2})[./](\d{1,2})[./](\d{4})',
                        r'(\d{1,2})[./](\d{1,2})[./](\d{2})',
                        r'(\d{4})-(\d{1,2})-(\d{1,2})'
                    ]
                    
                    for fmt in formats:
                        match = re.match(fmt, date_str)
                        if match:
                            parts = match.groups()
                            if len(parts[2]) == 2:
                                year = int(parts[2]) + 2000
                            else:
                                year = int(parts[2])
                            
                            if fmt.endswith(r'(\d{1,2})'):  # ISO format
                                return date(year, int(parts[1]), int(parts[0]))
                            else:  # DD/MM/YYYY format
                                return date(year, int(parts[1]), int(parts[0]))
                    
                    raise ValueError(f"Неверный формат даты: {date_str}")
            else:
                return value
                
        except Exception as e:
            raise ValueError(f"Ошибка преобразования значения '{value}' в тип {data_type} для поля {field_name}: {str(e)}")
    
    def _validate_row_data(self, row_data: Dict[str, Any], target_columns: List[ImportColumn]) -> List[str]:
        """Validate row data against target column requirements"""
        errors = []
        
        for target_col in target_columns:
            field_name = target_col.target_field
            value = row_data.get(field_name)
            
            # Check required fields
            if target_col.required and (value is None or value == ''):
                errors.append(f"Обязательное поле '{field_name}' не заполнено")
                continue
            
            # Apply validation rules if present
            if value is not None and target_col.validation_rules:
                for rule_name, rule_value in target_col.validation_rules.items():
                    if rule_name == 'min_value' and isinstance(value, (int, float)):
                        if value < rule_value:
                            errors.append(f"Значение поля '{field_name}' ({value}) меньше минимального ({rule_value})")
                    elif rule_name == 'max_value' and isinstance(value, (int, float)):
                        if value > rule_value:
                            errors.append(f"Значение поля '{field_name}' ({value}) больше максимального ({rule_value})")
                    elif rule_name == 'max_length' and isinstance(value, str):
                        if len(value) > rule_value:
                            errors.append(f"Длина поля '{field_name}' ({len(value)}) превышает максимальную ({rule_value})")
        
        return errors