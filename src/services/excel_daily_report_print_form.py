"""Excel daily report print form generator"""
from typing import Optional
from openpyxl.styles import Font, Alignment
from .excel_print_form_generator import ExcelPrintFormGenerator
from ..data.database_manager import DatabaseManager


class ExcelDailyReportPrintForm(ExcelPrintFormGenerator):
    """Generator for daily report print forms in Excel format"""
    
    TEMPLATE_NAME = "daily_report_template.xlsx"
    
    def __init__(self):
        """Initialize Excel daily report print form generator"""
        super().__init__()
        self.db = DatabaseManager().get_connection()
    
    def generate(self, report_id: int) -> Optional[bytes]:
        """
        Generate daily report print form in Excel format
        
        Args:
            report_id: ID of the daily report
            
        Returns:
            Excel content as bytes or None if report not found
        """
        # Load report data
        report_data = self._load_report_data(report_id)
        if not report_data:
            return None
        
        # Try to load template, otherwise create from scratch
        if self.template_exists(self.TEMPLATE_NAME):
            workbook = self.load_template(self.TEMPLATE_NAME)
            if workbook:
                self._fill_template(workbook, report_data)
            else:
                workbook = self._create_from_scratch(report_data)
        else:
            workbook = self._create_from_scratch(report_data)
        
        return self.save_to_bytes(workbook)
    
    def _load_report_data(self, report_id: int) -> Optional[dict]:
        """Load daily report data from database"""
        cursor = self.db.cursor()
        
        # Load report header
        cursor.execute("""
            SELECT 
                dr.id, dr.date,
                e.number as estimate_number, e.date as estimate_date,
                o.name as object_name,
                p.full_name as foreman_name
            FROM daily_reports dr
            LEFT JOIN estimates e ON dr.estimate_id = e.id
            LEFT JOIN objects o ON e.object_id = o.id
            LEFT JOIN persons p ON dr.foreman_id = p.id
            WHERE dr.id = ?
        """, (report_id,))
        
        row = cursor.fetchone()
        if not row:
            return None
        
        report_data = {
            'id': row['id'],
            'date': row['date'],
            'estimate_number': row['estimate_number'],
            'estimate_date': row['estimate_date'],
            'object_name': row['object_name'],
            'foreman_name': row['foreman_name'],
            'lines': []
        }
        
        # Load report lines
        cursor.execute("""
            SELECT 
                drl.line_number,
                w.name as work_name,
                drl.planned_labor,
                drl.actual_labor,
                drl.deviation_percent,
                drl.id as line_id,
                '' as unit,
                0 as quantity
            FROM daily_report_lines drl
            LEFT JOIN works w ON drl.work_id = w.id
            WHERE drl.daily_report_id = ?
            ORDER BY drl.line_number
        """, (report_id,))
        
        for line_row in cursor.fetchall():
            # Load executors
            cursor.execute("""
                SELECT p.full_name
                FROM daily_report_executors dre
                JOIN persons p ON dre.executor_id = p.id
                WHERE dre.report_line_id = ?
                ORDER BY p.full_name
            """, (line_row['line_id'],))
            
            executors = [row['full_name'] for row in cursor.fetchall()]
            
            report_data['lines'].append({
                'line_number': line_row['line_number'],
                'work_name': line_row['work_name'] or "",
                'unit': line_row['unit'] or "",
                'quantity': line_row['quantity'] or 0,
                'planned_labor': line_row['planned_labor'],
                'actual_labor': line_row['actual_labor'],
                'deviation_percent': line_row['deviation_percent'],
                'executors': executors
            })
        
        return report_data
    
    def _fill_template(self, workbook, report_data: dict):
        """Fill template with data"""
        sheet = workbook.active
        
        # Replace placeholders
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    value = str(cell.value)
                    value = value.replace('{DATE}', self.format_date(report_data['date']))
                    value = value.replace('{ESTIMATE_NUMBER}', report_data['estimate_number'] or '')
                    value = value.replace('{ESTIMATE_DATE}', self.format_date(report_data['estimate_date']))
                    value = value.replace('{OBJECT}', report_data['object_name'] or '')
                    value = value.replace('{FOREMAN}', report_data['foreman_name'] or '')
                    cell.value = value
        
        # Find table start
        table_start_row = None
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=50), start=1):
            for cell in row:
                if cell.value and str(cell.value).strip() in ['№\nп/п', '№ п/п', 'N', '№']:
                    table_start_row = row_idx + 1
                    break
            if table_start_row:
                break
        
        # Fill lines
        if table_start_row:
            self._fill_lines_in_template(sheet, report_data['lines'], table_start_row)
    
    def _fill_lines_in_template(self, sheet, lines: list, start_row: int):
        """Fill report lines in template"""
        current_row = start_row
        
        for line in lines:
            executors_str = ", ".join(line['executors']) if line['executors'] else ""
            
            self.set_cell_value(sheet, current_row, 1, line['line_number'])
            self.set_cell_value(sheet, current_row, 2, line['work_name'])
            self.set_cell_value(sheet, current_row, 3, line.get('unit', ''))
            self.set_cell_value(sheet, current_row, 4, float(line.get('quantity', 0)))
            self.set_cell_value(sheet, current_row, 5, float(line['planned_labor']))
            self.set_cell_value(sheet, current_row, 6, float(line['actual_labor']))
            self.set_cell_value(sheet, current_row, 7, float(line['deviation_percent']))
            self.set_cell_value(sheet, current_row, 8, executors_str)
            
            current_row += 1
    
    def _create_from_scratch(self, report_data: dict):
        """Create Excel document from scratch"""
        workbook = self.create_workbook()
        sheet = workbook.active
        sheet.title = "Ежедневный отчет"
        
        current_row = 1
        
        # Title
        self.set_cell_value(sheet, current_row, 1, "ЕЖЕДНЕВНЫЙ ОТЧЕТ О ВЫПОЛНЕННЫХ РАБОТАХ")
        self.merge_cells(sheet, current_row, 1, current_row, 8)
        title_style = self.create_title_style()
        self.set_cell_style(sheet, current_row, 1, **title_style)
        current_row += 2
        
        # Info section
        self.set_cell_value(sheet, current_row, 1, "Дата отчета:")
        self.set_cell_value(sheet, current_row, 2, self.format_date(report_data['date']))
        current_row += 1
        
        estimate_info = f"{report_data['estimate_number']} от {self.format_date(report_data['estimate_date'])}"
        self.set_cell_value(sheet, current_row, 1, "Смета №:")
        self.set_cell_value(sheet, current_row, 2, estimate_info)
        current_row += 1
        
        self.set_cell_value(sheet, current_row, 1, "Объект:")
        self.set_cell_value(sheet, current_row, 2, report_data['object_name'] or "")
        current_row += 1
        
        self.set_cell_value(sheet, current_row, 1, "Бригадир:")
        self.set_cell_value(sheet, current_row, 2, report_data['foreman_name'] or "")
        current_row += 2
        
        # Table header
        header_row = current_row
        headers = ["№\nп/п", "Наименование\nвыполненных работ", "Ед.\nизм.", "Объем\nработ",
                  "Плановые\nтрудозатраты, ч", "Фактические\nтрудозатраты, ч",
                  "Отклонение,\n%", "Исполнители\n(ФИО)"]
        
        for col, header in enumerate(headers, start=1):
            self.set_cell_value(sheet, header_row, col, header)
            header_style = self.create_header_style()
            self.set_cell_style(sheet, header_row, col, **header_style)
        current_row += 1
        
        # Set column widths
        widths = [5, 30, 8, 10, 12, 12, 10, 25]
        for col, width in enumerate(widths, start=1):
            self.set_column_width(sheet, col, width)
        
        # Data rows
        data_style = self.create_data_style()
        number_style = self.create_number_style()
        
        for line in report_data['lines']:
            executors_str = ", ".join(line['executors']) if line['executors'] else ""
            
            self.set_cell_value(sheet, current_row, 1, line['line_number'])
            self.set_cell_value(sheet, current_row, 2, line['work_name'])
            self.set_cell_value(sheet, current_row, 3, line.get('unit', ''))
            self.set_cell_value(sheet, current_row, 4, float(line.get('quantity', 0)))
            self.set_cell_value(sheet, current_row, 5, float(line['planned_labor']))
            self.set_cell_value(sheet, current_row, 6, float(line['actual_labor']))
            self.set_cell_value(sheet, current_row, 7, float(line['deviation_percent']))
            self.set_cell_value(sheet, current_row, 8, executors_str)
            
            # Apply styles
            for col in range(1, 9):
                if col in [1, 2, 3, 8]:
                    self.set_cell_style(sheet, current_row, col, **data_style)
                else:
                    self.set_cell_style(sheet, current_row, col, **number_style)
            
            current_row += 1
        
        # Totals
        current_row += 1
        total_planned = sum(line['planned_labor'] for line in report_data['lines'])
        total_actual = sum(line['actual_labor'] for line in report_data['lines'])
        total_deviation = ((total_actual - total_planned) / total_planned * 100) if total_planned > 0 else 0
        
        self.set_cell_value(sheet, current_row, 1, "ИТОГО:")
        self.set_cell_style(sheet, current_row, 1, font=Font(bold=True, size=11))
        current_row += 1
        
        self.set_cell_value(sheet, current_row, 1, f"Плановые трудозатраты: {self.format_number(total_planned, 2)} ч.")
        current_row += 1
        
        self.set_cell_value(sheet, current_row, 1, f"Фактические трудозатраты: {self.format_number(total_actual, 2)} ч.")
        current_row += 1
        
        self.set_cell_value(sheet, current_row, 1, f"Отклонение: {self.format_number(total_deviation, 1)}%")
        current_row += 2
        
        # Signatures
        foreman_name = report_data['foreman_name'] or "_______________"
        self.set_cell_value(sheet, current_row, 1, "Бригадир:")
        self.set_cell_value(sheet, current_row, 2, foreman_name)
        self.set_cell_value(sheet, current_row, 4, "____________________")
        current_row += 1
        self.set_cell_value(sheet, current_row, 4, "(подпись)")
        current_row += 2
        
        self.set_cell_value(sheet, current_row, 1, "Принял:")
        self.set_cell_value(sheet, current_row, 2, "_______________")
        self.set_cell_value(sheet, current_row, 4, "____________________")
        current_row += 1
        self.set_cell_value(sheet, current_row, 4, "(подпись)")
        
        return workbook
    
    def create_template(self) -> bool:
        """Create default template file"""
        try:
            workbook = self.create_workbook()
            sheet = workbook.active
            sheet.title = "Ежедневный отчет"
            
            # Title
            self.set_cell_value(sheet, 1, 1, "ЕЖЕДНЕВНЫЙ ОТЧЕТ О ВЫПОЛНЕННЫХ РАБОТАХ")
            self.merge_cells(sheet, 1, 1, 1, 8)
            title_style = self.create_title_style()
            self.set_cell_style(sheet, 1, 1, **title_style)
            
            # Info with placeholders
            self.set_cell_value(sheet, 3, 1, "Дата отчета:")
            self.set_cell_value(sheet, 3, 2, "{DATE}")
            self.set_cell_value(sheet, 4, 1, "Смета №:")
            self.set_cell_value(sheet, 4, 2, "{ESTIMATE_NUMBER} от {ESTIMATE_DATE}")
            self.set_cell_value(sheet, 5, 1, "Объект:")
            self.set_cell_value(sheet, 5, 2, "{OBJECT}")
            self.set_cell_value(sheet, 6, 1, "Бригадир:")
            self.set_cell_value(sheet, 6, 2, "{FOREMAN}")
            
            # Table header
            header_row = 8
            headers = ["№\nп/п", "Наименование\nвыполненных работ", "Ед.\nизм.", "Объем\nработ",
                      "Плановые\nтрудозатраты, ч", "Фактические\nтрудозатраты, ч",
                      "Отклонение,\n%", "Исполнители\n(ФИО)"]
            
            for col, header in enumerate(headers, start=1):
                self.set_cell_value(sheet, header_row, col, header)
                header_style = self.create_header_style()
                self.set_cell_style(sheet, header_row, col, **header_style)
            
            # Set column widths
            widths = [5, 30, 8, 10, 12, 12, 10, 25]
            for col, width in enumerate(widths, start=1):
                self.set_column_width(sheet, col, width)
            
            # Save template
            template_path = self.get_template_path(self.TEMPLATE_NAME)
            workbook.save(template_path)
            return True
        except Exception as e:
            print(f"Error creating template: {e}")
            return False
