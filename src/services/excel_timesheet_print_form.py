"""Excel timesheet print form generator"""
from typing import Optional
from openpyxl.styles import Font, Alignment, Border, Side
from calendar import monthrange
from datetime import datetime
from .excel_print_form_generator import ExcelPrintFormGenerator
from ..data.database_manager import DatabaseManager


class ExcelTimesheetPrintForm(ExcelPrintFormGenerator):
    """Generator for timesheet print forms in Excel format"""
    
    TEMPLATE_NAME = "timesheet_template.xlsx"
    
    def __init__(self):
        """Initialize Excel timesheet print form generator"""
        super().__init__()
        self.db = DatabaseManager().get_connection()
    
    def generate(self, timesheet_id: int) -> Optional[bytes]:
        """
        Generate timesheet print form in Excel format
        
        Args:
            timesheet_id: ID of the timesheet
            
        Returns:
            Excel content as bytes or None if timesheet not found
        """
        # Load timesheet data
        timesheet_data = self._load_timesheet_data(timesheet_id)
        if not timesheet_data:
            return None
        
        # Try to load template, otherwise create from scratch
        if self.template_exists(self.TEMPLATE_NAME):
            workbook = self.load_template(self.TEMPLATE_NAME)
            if workbook:
                self._fill_template(workbook, timesheet_data)
            else:
                workbook = self._create_from_scratch(timesheet_data)
        else:
            workbook = self._create_from_scratch(timesheet_data)
        
        return self.save_to_bytes(workbook)
    
    def _load_timesheet_data(self, timesheet_id: int) -> Optional[dict]:
        """Load timesheet data from database"""
        cursor = self.db.cursor()
        
        # Load timesheet header
        cursor.execute("""
            SELECT 
                t.id, t.number, t.date, t.month_year,
                o.name as object_name,
                e.number as estimate_number,
                p.full_name as foreman_name
            FROM timesheets t
            LEFT JOIN objects o ON t.object_id = o.id
            LEFT JOIN estimates e ON t.estimate_id = e.id
            LEFT JOIN persons p ON t.foreman_id = p.id
            WHERE t.id = ?
        """, (timesheet_id,))
        
        row = cursor.fetchone()
        if not row:
            return None
        
        # Parse month_year to get days in month
        year, month = map(int, row['month_year'].split('-'))
        days_in_month = monthrange(year, month)[1]
        
        timesheet_data = {
            'id': row['id'],
            'number': row['number'],
            'date': row['date'],
            'month_year': row['month_year'],
            'year': year,
            'month': month,
            'days_in_month': days_in_month,
            'object_name': row['object_name'],
            'estimate_number': row['estimate_number'],
            'foreman_name': row['foreman_name'],
            'lines': []
        }
        
        # Load timesheet lines
        cursor.execute("""
            SELECT 
                tl.line_number,
                tl.employee_id,
                p.full_name as employee_name,
                tl.hourly_rate,
                tl.day_01, tl.day_02, tl.day_03, tl.day_04, tl.day_05,
                tl.day_06, tl.day_07, tl.day_08, tl.day_09, tl.day_10,
                tl.day_11, tl.day_12, tl.day_13, tl.day_14, tl.day_15,
                tl.day_16, tl.day_17, tl.day_18, tl.day_19, tl.day_20,
                tl.day_21, tl.day_22, tl.day_23, tl.day_24, tl.day_25,
                tl.day_26, tl.day_27, tl.day_28, tl.day_29, tl.day_30,
                tl.day_31,
                tl.total_hours,
                tl.total_amount
            FROM timesheet_lines tl
            LEFT JOIN persons p ON tl.employee_id = p.id
            WHERE tl.timesheet_id = ?
            ORDER BY tl.line_number
        """, (timesheet_id,))
        
        for line_row in cursor.fetchall():
            # Collect day values
            days = {}
            for day in range(1, days_in_month + 1):
                day_col = f'day_{day:02d}'
                hours = line_row[day_col] if day_col in line_row.keys() else 0
                if hours > 0:
                    days[day] = hours
            
            timesheet_data['lines'].append({
                'line_number': line_row['line_number'],
                'employee_name': line_row['employee_name'] or "",
                'hourly_rate': line_row['hourly_rate'],
                'days': days,
                'total_hours': line_row['total_hours'],
                'total_amount': line_row['total_amount']
            })
        
        return timesheet_data
    
    def _fill_template(self, workbook, timesheet_data: dict):
        """Fill template with data"""
        sheet = workbook.active
        
        # Replace placeholders
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    value = str(cell.value)
                    value = value.replace('{NUMBER}', timesheet_data['number'] or '')
                    value = value.replace('{DATE}', self.format_date(timesheet_data['date']))
                    value = value.replace('{MONTH_YEAR}', timesheet_data['month_year'])
                    value = value.replace('{OBJECT}', timesheet_data['object_name'] or '')
                    value = value.replace('{ESTIMATE}', timesheet_data['estimate_number'] or '')
                    value = value.replace('{FOREMAN}', timesheet_data['foreman_name'] or '')
                    cell.value = value
        
        # Find table start
        table_start_row = None
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=50), start=1):
            for cell in row:
                if cell.value and str(cell.value).strip() in ['№\nп/п', '№ п/п', 'N', '№']:
                    table_start_row = row_idx + 1
                    break
            if table_start_row:
                break
        
        # Fill lines
        if table_start_row:
            self._fill_lines_in_template(sheet, timesheet_data, table_start_row)
    
    def _fill_lines_in_template(self, sheet, timesheet_data: dict, start_row: int):
        """Fill timesheet lines in template"""
        current_row = start_row
        days_in_month = timesheet_data['days_in_month']
        
        for line in timesheet_data['lines']:
            # Line number
            self.set_cell_value(sheet, current_row, 1, line['line_number'])
            
            # Employee name
            self.set_cell_value(sheet, current_row, 2, line['employee_name'])
            
            # Hourly rate
            self.set_cell_value(sheet, current_row, 3, self.safe_float(line['hourly_rate'] or 0))
            
            # Day columns
            for day in range(1, days_in_month + 1):
                col = 3 + day
                hours = line['days'].get(day, 0)
                if hours > 0:
                    self.set_cell_value(sheet, current_row, col, self.safe_float(hours or 0))
                else:
                    self.set_cell_value(sheet, current_row, col, "")
            
            # Total hours
            total_col = 3 + days_in_month + 1
            self.set_cell_value(sheet, current_row, total_col, self.safe_float(line['total_hours'] or 0))
            
            # Total amount
            amount_col = 3 + days_in_month + 2
            self.set_cell_value(sheet, current_row, amount_col, self.safe_float(line['total_amount'] or 0))
            
            current_row += 1
    
    def _create_from_scratch(self, timesheet_data: dict):
        """Create Excel document from scratch"""
        workbook = self.create_workbook()
        sheet = workbook.active
        sheet.title = "Табель"
        
        current_row = 1
        days_in_month = timesheet_data['days_in_month']
        year = timesheet_data['year']
        month = timesheet_data['month']
        
        # Title
        self.set_cell_value(sheet, current_row, 1, f"ТАБЕЛЬ УЧЕТА РАБОЧЕГО ВРЕМЕНИ №{timesheet_data['number']}")
        num_cols = 3 + days_in_month + 2  # Employee, Rate, Days, Total, Amount
        self.merge_cells(sheet, current_row, 1, current_row, num_cols)
        title_style = self.create_title_style()
        self.set_cell_style(sheet, current_row, 1, **title_style)
        current_row += 2
        
        # Info section
        self.set_cell_value(sheet, current_row, 1, "Дата:")
        self.set_cell_value(sheet, current_row, 2, self.format_date(timesheet_data['date']))
        current_row += 1
        
        self.set_cell_value(sheet, current_row, 1, "Период:")
        month_names = ['', 'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                      'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
        period_text = f"{month_names[month]} {year}"
        self.set_cell_value(sheet, current_row, 2, period_text)
        current_row += 1
        
        self.set_cell_value(sheet, current_row, 1, "Объект:")
        self.set_cell_value(sheet, current_row, 2, timesheet_data['object_name'] or "")
        current_row += 1
        
        self.set_cell_value(sheet, current_row, 1, "Смета №:")
        self.set_cell_value(sheet, current_row, 2, timesheet_data['estimate_number'] or "")
        current_row += 1
        
        self.set_cell_value(sheet, current_row, 1, "Бригадир:")
        self.set_cell_value(sheet, current_row, 2, timesheet_data['foreman_name'] or "")
        current_row += 2
        
        # Table header - Row 1
        header_row1 = current_row
        self.set_cell_value(sheet, header_row1, 1, "№\nп/п")
        self.set_cell_value(sheet, header_row1, 2, "Сотрудник")
        self.set_cell_value(sheet, header_row1, 3, "Ставка,\nруб/ч")
        
        # Days header
        for day in range(1, days_in_month + 1):
            col = 3 + day
            self.set_cell_value(sheet, header_row1, col, str(day))
        
        self.set_cell_value(sheet, header_row1, 3 + days_in_month + 1, "Итого,\nч")
        self.set_cell_value(sheet, header_row1, 3 + days_in_month + 2, "Сумма,\nруб")
        
        # Apply header style
        header_style = self.create_header_style()
        for col in range(1, num_cols + 1):
            self.set_cell_style(sheet, header_row1, col, **header_style)
        
        current_row += 1
        
        # Set column widths
        self.set_column_width(sheet, 1, 5)   # № п/п
        self.set_column_width(sheet, 2, 25)  # Сотрудник
        self.set_column_width(sheet, 3, 10)  # Ставка
        
        # Day columns
        for day in range(1, days_in_month + 1):
            col = 3 + day
            self.set_column_width(sheet, col, 4)
        
        self.set_column_width(sheet, 3 + days_in_month + 1, 10)  # Итого
        self.set_column_width(sheet, 3 + days_in_month + 2, 12)  # Сумма
        
        # Data rows
        data_style = self.create_data_style()
        number_style = self.create_number_style()
        
        for line in timesheet_data['lines']:
            # Line number
            self.set_cell_value(sheet, current_row, 1, line['line_number'])
            self.set_cell_style(sheet, current_row, 1, **data_style)
            
            # Employee name
            self.set_cell_value(sheet, current_row, 2, line['employee_name'])
            self.set_cell_style(sheet, current_row, 2, **data_style)
            
            # Hourly rate
            self.set_cell_value(sheet, current_row, 3, self.safe_float(line['hourly_rate'] or 0))
            self.set_cell_style(sheet, current_row, 3, **number_style)
            
            # Day columns
            for day in range(1, days_in_month + 1):
                col = 3 + day
                hours = line['days'].get(day, 0)
                if hours > 0:
                    self.set_cell_value(sheet, current_row, col, self.safe_float(hours or 0))
                else:
                    self.set_cell_value(sheet, current_row, col, "")
                self.set_cell_style(sheet, current_row, col, **number_style)
            
            # Total hours
            total_col = 3 + days_in_month + 1
            self.set_cell_value(sheet, current_row, total_col, self.safe_float(line['total_hours'] or 0))
            self.set_cell_style(sheet, current_row, total_col, **number_style)
            
            # Total amount
            amount_col = 3 + days_in_month + 2
            self.set_cell_value(sheet, current_row, amount_col, self.safe_float(line['total_amount'] or 0))
            self.set_cell_style(sheet, current_row, amount_col, **number_style)
            
            current_row += 1
        
        # Totals row
        current_row += 1
        self.set_cell_value(sheet, current_row, 1, "ИТОГО:")
        self.merge_cells(sheet, current_row, 1, current_row, 3)
        self.set_cell_style(sheet, current_row, 1, font=Font(bold=True, size=11))
        
        # Calculate totals
        total_hours = sum(line['total_hours'] for line in timesheet_data['lines'])
        total_amount = sum(line['total_amount'] for line in timesheet_data['lines'])
        
        total_col = 3 + days_in_month + 1
        amount_col = 3 + days_in_month + 2
        
        self.set_cell_value(sheet, current_row, total_col, self.safe_float(total_hours or 0))
        self.set_cell_style(sheet, current_row, total_col, 
                           font=Font(bold=True, size=11),
                           alignment=Alignment(horizontal='right', vertical='center'),
                           border=Border(
                               left=Side(style='thin'),
                               right=Side(style='thin'),
                               top=Side(style='thin'),
                               bottom=Side(style='thin')
                           ))
        
        self.set_cell_value(sheet, current_row, amount_col, self.safe_float(total_amount or 0))
        self.set_cell_style(sheet, current_row, amount_col,
                           font=Font(bold=True, size=11),
                           alignment=Alignment(horizontal='right', vertical='center'),
                           border=Border(
                               left=Side(style='thin'),
                               right=Side(style='thin'),
                               top=Side(style='thin'),
                               bottom=Side(style='thin')
                           ))
        
        current_row += 2
        
        # Signatures
        self.set_cell_value(sheet, current_row, 1, "Бригадир:")
        self.set_cell_value(sheet, current_row, 2, timesheet_data['foreman_name'] or "_______________")
        self.set_cell_value(sheet, current_row, 4, "____________________")
        current_row += 1
        self.set_cell_value(sheet, current_row, 4, "(подпись)")
        
        return workbook
    
    def create_template(self) -> bool:
        """Create default template file"""
        try:
            workbook = self.create_workbook()
            sheet = workbook.active
            sheet.title = "Табель"
            
            # Title
            self.set_cell_value(sheet, 1, 1, "ТАБЕЛЬ УЧЕТА РАБОЧЕГО ВРЕМЕНИ №{NUMBER}")
            self.merge_cells(sheet, 1, 1, 1, 35)  # Approximate column count
            title_style = self.create_title_style()
            self.set_cell_style(sheet, 1, 1, **title_style)
            
            # Info with placeholders
            self.set_cell_value(sheet, 3, 1, "Дата:")
            self.set_cell_value(sheet, 3, 2, "{DATE}")
            self.set_cell_value(sheet, 4, 1, "Период:")
            self.set_cell_value(sheet, 4, 2, "{MONTH_YEAR}")
            self.set_cell_value(sheet, 5, 1, "Объект:")
            self.set_cell_value(sheet, 5, 2, "{OBJECT}")
            self.set_cell_value(sheet, 6, 1, "Смета №:")
            self.set_cell_value(sheet, 6, 2, "{ESTIMATE}")
            self.set_cell_value(sheet, 7, 1, "Бригадир:")
            self.set_cell_value(sheet, 7, 2, "{FOREMAN}")
            
            # Table header
            header_row = 9
            self.set_cell_value(sheet, header_row, 1, "№\nп/п")
            self.set_cell_value(sheet, header_row, 2, "Сотрудник")
            self.set_cell_value(sheet, header_row, 3, "Ставка,\nруб/ч")
            
            # Days 1-31
            for day in range(1, 32):
                col = 3 + day
                self.set_cell_value(sheet, header_row, col, str(day))
            
            self.set_cell_value(sheet, header_row, 35, "Итого,\nч")
            self.set_cell_value(sheet, header_row, 36, "Сумма,\nруб")
            
            # Apply header style
            header_style = self.create_header_style()
            for col in range(1, 37):
                self.set_cell_style(sheet, header_row, col, **header_style)
            
            # Set column widths
            self.set_column_width(sheet, 1, 5)
            self.set_column_width(sheet, 2, 25)
            self.set_column_width(sheet, 3, 10)
            for day in range(1, 32):
                self.set_column_width(sheet, 3 + day, 4)
            self.set_column_width(sheet, 35, 10)
            self.set_column_width(sheet, 36, 12)
            
            # Save template
            template_path = self.get_template_path(self.TEMPLATE_NAME)
            workbook.save(template_path)
            return True
        except Exception as e:
            print(f"Error creating template: {e}")
            return False
