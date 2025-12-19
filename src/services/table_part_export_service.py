"""
Table part export service for handling Excel and CSV exports.

This service provides functionality to export table part data to
Excel (.xlsx) and CSV files with format selection and customization.
"""

import openpyxl
import csv
import os
from typing import Dict, List, Optional, Any, Union
from dataclasses import dataclass
from enum import Enum
from datetime import datetime, date
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class ExportFormat(Enum):
    """Supported export file formats"""
    EXCEL = "excel"
    CSV = "csv"


class ExportValidationError(Exception):
    """Exception raised when export validation fails"""
    pass


@dataclass
class ExportColumn:
    """Configuration for an export column"""
    field_name: str  # Source field name in data
    display_name: str  # Column header in export file
    data_type: str  # Data type (str, int, float, date, bool)
    width: Optional[int] = None  # Column width for Excel
    format_string: Optional[str] = None  # Format string for values
    include_in_export: bool = True


@dataclass
class ExportOptions:
    """Options for export operation"""
    include_headers: bool = True
    date_format: str = "dd.mm.yyyy"
    number_format: str = "0.00"
    encoding: str = "utf-8"  # For CSV exports
    delimiter: str = ","  # For CSV exports
    apply_formatting: bool = True  # For Excel exports
    auto_fit_columns: bool = True  # For Excel exports


@dataclass
class ExportResult:
    """Result of export operation"""
    success: bool
    file_path: str
    exported_rows: int
    errors: List[str]
    warnings: List[str]


class TablePartExportService:
    """Service for exporting table part data"""
    
    def __init__(self):
        self.supported_formats = [ExportFormat.EXCEL, ExportFormat.CSV]
        self.max_export_rows = 100000
    
    def export_data(
        self,
        data: List[Dict[str, Any]],
        file_path: str,
        columns: List[ExportColumn],
        options: Optional[ExportOptions] = None
    ) -> ExportResult:
        """
        Export data to specified file format.
        
        Args:
            data: List of data dictionaries to export
            file_path: Target file path
            columns: Column definitions for export
            options: Export options
            
        Returns:
            ExportResult with success status and details
        """
        if options is None:
            options = ExportOptions()
        
        # Validate data
        if not data:
            return ExportResult(
                success=False,
                file_path=file_path,
                exported_rows=0,
                errors=["Нет данных для экспорта"],
                warnings=[]
            )
        
        if len(data) > self.max_export_rows:
            return ExportResult(
                success=False,
                file_path=file_path,
                exported_rows=0,
                errors=[f"Слишком много строк для экспорта. Максимум: {self.max_export_rows}"],
                warnings=[]
            )
        
        # Detect format from file extension
        format_type = self._detect_format(file_path)
        
        # Filter columns to export
        export_columns = [col for col in columns if col.include_in_export]
        
        if not export_columns:
            return ExportResult(
                success=False,
                file_path=file_path,
                exported_rows=0,
                errors=["Не выбрано ни одной колонки для экспорта"],
                warnings=[]
            )
        
        # Export based on format
        if format_type == ExportFormat.EXCEL:
            return self._export_to_excel(data, file_path, export_columns, options)
        elif format_type == ExportFormat.CSV:
            return self._export_to_csv(data, file_path, export_columns, options)
        else:
            return ExportResult(
                success=False,
                file_path=file_path,
                exported_rows=0,
                errors=[f"Неподдерживаемый формат файла: {format_type}"],
                warnings=[]
            )
    
    def _detect_format(self, file_path: str) -> ExportFormat:
        """Detect export format from file extension"""
        ext = os.path.splitext(file_path)[1].lower()
        if ext == '.xlsx':
            return ExportFormat.EXCEL
        elif ext == '.csv':
            return ExportFormat.CSV
        else:
            raise ExportValidationError(f"Неподдерживаемый формат файла: {ext}")
    
    def _export_to_excel(
        self,
        data: List[Dict[str, Any]],
        file_path: str,
        columns: List[ExportColumn],
        options: ExportOptions
    ) -> ExportResult:
        """Export data to Excel file"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Экспорт данных"
            
            errors = []
            warnings = []
            exported_rows = 0
            
            # Write headers
            if options.include_headers:
                for col_idx, column in enumerate(columns, 1):
                    cell = ws.cell(row=1, column=col_idx, value=column.display_name)
                    
                    if options.apply_formatting:
                        # Header formatting
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = Border(
                            left=Side(style="thin"),
                            right=Side(style="thin"),
                            top=Side(style="thin"),
                            bottom=Side(style="thin")
                        )
                
                exported_rows = 1
            
            # Write data rows
            for row_idx, row_data in enumerate(data, start=2 if options.include_headers else 1):
                try:
                    for col_idx, column in enumerate(columns, 1):
                        value = row_data.get(column.field_name)
                        
                        # Convert and format value
                        formatted_value = self._format_value(value, column, options)
                        
                        cell = ws.cell(row=row_idx, column=col_idx, value=formatted_value)
                        
                        if options.apply_formatting:
                            # Data formatting
                            cell.alignment = Alignment(vertical="center")
                            cell.border = Border(
                                left=Side(style="thin"),
                                right=Side(style="thin"),
                                top=Side(style="thin"),
                                bottom=Side(style="thin")
                            )
                            
                            # Apply data type specific formatting
                            if column.data_type in ['int', 'float']:
                                cell.alignment = Alignment(horizontal="right", vertical="center")
                                if column.format_string:
                                    cell.number_format = column.format_string
                                elif column.data_type == 'float':
                                    cell.number_format = options.number_format
                            elif column.data_type == 'date':
                                cell.number_format = options.date_format
                    
                    exported_rows += 1
                    
                except Exception as e:
                    errors.append(f"Ошибка в строке {row_idx}: {str(e)}")
            
            # Set column widths
            if options.auto_fit_columns:
                for col_idx, column in enumerate(columns, 1):
                    col_letter = get_column_letter(col_idx)
                    
                    if column.width:
                        ws.column_dimensions[col_letter].width = column.width
                    else:
                        # Auto-fit based on content
                        max_length = len(column.display_name)
                        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                            for cell in row:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                        
                        # Set width with some padding
                        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
            
            # Save workbook
            wb.save(file_path)
            
            return ExportResult(
                success=True,
                file_path=file_path,
                exported_rows=exported_rows - (1 if options.include_headers else 0),
                errors=errors,
                warnings=warnings
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                file_path=file_path,
                exported_rows=0,
                errors=[f"Ошибка экспорта в Excel: {str(e)}"],
                warnings=[]
            )
    
    def _export_to_csv(
        self,
        data: List[Dict[str, Any]],
        file_path: str,
        columns: List[ExportColumn],
        options: ExportOptions
    ) -> ExportResult:
        """Export data to CSV file"""
        try:
            errors = []
            warnings = []
            exported_rows = 0
            
            with open(file_path, 'w', encoding=options.encoding, newline='') as file:
                writer = csv.writer(file, delimiter=options.delimiter)
                
                # Write headers
                if options.include_headers:
                    headers = [column.display_name for column in columns]
                    writer.writerow(headers)
                    exported_rows = 1
                
                # Write data rows
                for row_idx, row_data in enumerate(data, start=1):
                    try:
                        row_values = []
                        for column in columns:
                            value = row_data.get(column.field_name)
                            formatted_value = self._format_value(value, column, options)
                            
                            # Convert to string for CSV
                            if formatted_value is None:
                                row_values.append("")
                            else:
                                row_values.append(str(formatted_value))
                        
                        writer.writerow(row_values)
                        exported_rows += 1
                        
                    except Exception as e:
                        errors.append(f"Ошибка в строке {row_idx}: {str(e)}")
            
            return ExportResult(
                success=True,
                file_path=file_path,
                exported_rows=exported_rows - (1 if options.include_headers else 0),
                errors=errors,
                warnings=warnings
            )
            
        except Exception as e:
            return ExportResult(
                success=False,
                file_path=file_path,
                exported_rows=0,
                errors=[f"Ошибка экспорта в CSV: {str(e)}"],
                warnings=[]
            )
    
    def _format_value(self, value: Any, column: ExportColumn, options: ExportOptions) -> Any:
        """Format value according to column type and options"""
        if value is None:
            return None
        
        try:
            if column.data_type == 'str':
                return str(value)
            elif column.data_type == 'int':
                if isinstance(value, (int, float)):
                    return int(value)
                return int(float(str(value).replace(',', '.')))
            elif column.data_type == 'float':
                if isinstance(value, (int, float)):
                    return float(value)
                return float(str(value).replace(',', '.'))
            elif column.data_type == 'bool':
                if isinstance(value, bool):
                    return "Да" if value else "Нет"
                str_val = str(value).lower().strip()
                return "Да" if str_val in ['true', '1', 'да', 'yes', 'истина'] else "Нет"
            elif column.data_type == 'date':
                if isinstance(value, date):
                    return value
                elif isinstance(value, datetime):
                    return value.date()
                else:
                    # Try to parse date string
                    import re
                    date_str = str(value).strip()
                    
                    # Try different date formats
                    formats = [
                        r'(\d{1,2})[./](\d{1,2})[./](\d{4})',
                        r'(\d{1,2})[./](\d{1,2})[./](\d{2})',
                        r'(\d{4})-(\d{1,2})-(\d{1,2})'
                    ]
                    
                    for fmt in formats:
                        match = re.match(fmt, date_str)
                        if match:
                            parts = match.groups()
                            if len(parts[2]) == 2:
                                year = int(parts[2]) + 2000
                            else:
                                year = int(parts[2])
                            
                            if fmt.endswith(r'(\d{1,2})'):  # ISO format
                                return date(year, int(parts[1]), int(parts[0]))
                            else:  # DD/MM/YYYY format
                                return date(year, int(parts[1]), int(parts[0]))
                    
                    return str(value)  # Return as string if can't parse
            else:
                return value
                
        except Exception:
            # Return original value if conversion fails
            return value
    
    def get_suggested_filename(self, base_name: str, format_type: ExportFormat) -> str:
        """Get suggested filename with timestamp"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if format_type == ExportFormat.EXCEL:
            return f"{base_name}_{timestamp}.xlsx"
        elif format_type == ExportFormat.CSV:
            return f"{base_name}_{timestamp}.csv"
        else:
            return f"{base_name}_{timestamp}.txt"
    
    def validate_file_path(self, file_path: str) -> bool:
        """Validate that file path is writable"""
        try:
            # Check if directory exists and is writable
            directory = os.path.dirname(file_path)
            if directory and not os.path.exists(directory):
                os.makedirs(directory, exist_ok=True)
            
            # Try to create a test file
            test_file = file_path + ".test"
            with open(test_file, 'w') as f:
                f.write("test")
            os.remove(test_file)
            
            return True
            
        except Exception:
            return False