"""Excel daily report import service"""
import openpyxl
from datetime import date
from typing import Optional, Tuple, List
from ..data.models.daily_report import DailyReport, DailyReportLine
from ..data.database_manager import DatabaseManager


class ExcelDailyReportImportService:
    def __init__(self):
        self.db = DatabaseManager().get_connection()
    
    def import_daily_report(self, file_path: str) -> Tuple[Optional[DailyReport], str]:
        """
        Import daily report from Excel file
        Returns: (DailyReport object or None, error message)
        """
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)  # data_only=True to get calculated values
            ws = wb.active
            
            daily_report = DailyReport(
                date=date.today()
            )
            
            # Parse header information
            estimate_id = None
            foreman_id = None
            
            for row in ws.iter_rows(min_row=1, max_row=15):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        # Extract estimate number/reference
                        if "смета" in cell.value.lower() or "№" in cell.value:
                            # Try to extract estimate reference
                            estimate_ref = self._extract_estimate_reference(cell.value)
                            if estimate_ref:
                                estimate_id = self._find_estimate_by_number(estimate_ref)
                        
                        # Extract foreman name
                        elif "бригадир" in cell.value.lower() or "прораб" in cell.value.lower():
                            foreman_name = self._extract_foreman_name(cell.value)
                            if foreman_name:
                                foreman_id = self._find_or_create_person(foreman_name)
                        
                        # Extract date if present
                        elif "дата" in cell.value.lower():
                            report_date = self._extract_date(cell.value)
                            if report_date:
                                daily_report.date = report_date
            
            # Update the daily report with found IDs
            if estimate_id:
                daily_report.estimate_id = estimate_id
            if foreman_id:
                daily_report.foreman_id = foreman_id
            
            # Find the header row (contains work names or codes)
            header_row = None
            work_name_col = None
            quantity_col = None
            
            # Search in a wider range and be more flexible with keywords
            for row_idx in range(1, min(51, ws.max_row + 1)):
                row_has_work_header = False
                row_has_quantity_header = False
                temp_work_col = None
                temp_quantity_col = None
                
                for col_idx in range(1, min(21, ws.max_column + 1)):
                    cell_value = self._get_cell_value(ws, row_idx, col_idx)
                    if cell_value and isinstance(cell_value, str):
                        cell_value = cell_value.lower().strip()
                        
                        # More flexible work column detection
                        work_keywords = ["наименование", "работ", "код", "шифр", "позиция", "вид работ", "описание"]
                        if any(keyword in cell_value for keyword in work_keywords):
                            row_has_work_header = True
                            temp_work_col = col_idx - 1  # Convert to 0-based for consistency
                        
                        # More flexible quantity column detection  
                        quantity_keywords = ["количество", "объем", "факт", "выполнено", "кол-во", "объём", "план", "норма"]
                        if any(keyword in cell_value for keyword in quantity_keywords):
                            row_has_quantity_header = True
                            temp_quantity_col = col_idx - 1  # Convert to 0-based for consistency
                
                # If we found both headers in this row, use it
                if row_has_work_header and row_has_quantity_header:
                    header_row = row_idx
                    work_name_col = temp_work_col
                    quantity_col = temp_quantity_col
                    break
            
            if not header_row:
                # Debug: collect all text values to help diagnose the issue
                debug_info = []
                for row_idx in range(1, min(51, ws.max_row + 1)):
                    row_texts = []
                    for col_idx in range(1, min(21, ws.max_column + 1)):  # Check first 20 columns
                        cell_value = self._get_cell_value(ws, row_idx, col_idx)
                        if cell_value and isinstance(cell_value, str):
                            row_texts.append(f"'{cell_value.strip()}'")
                    if row_texts:
                        debug_info.append(f"Строка {row_idx}: {', '.join(row_texts)}")
                
                debug_message = "\n".join(debug_info[:10])  # Show first 10 rows with text
                return None, f"Не найден заголовок таблицы с работами. Убедитесь, что в файле есть строка с заголовками колонок, содержащими слова 'наименование', 'работ', 'код' для работ и 'количество', 'объем', 'факт' для количества.\n\nОбнаруженный текст в файле:\n{debug_message}"
            
            if work_name_col is None:
                return None, "Не найдена колонка с наименованием работ. Ожидаются заголовки: 'наименование', 'работ', 'код', 'шифр', 'позиция'"
            
            if quantity_col is None:
                return None, "Не найдена колонка с количеством. Ожидаются заголовки: 'количество', 'объем', 'факт', 'выполнено', 'кол-во'"
            
            # Parse table rows
            line_number = 0
            for row_idx in range(header_row + 1, ws.max_row + 1):
                # Check if row has any content
                has_content = False
                for col_idx in range(1, ws.max_column + 1):
                    if self._get_cell_value(ws, row_idx, col_idx):
                        has_content = True
                        break
                
                if not has_content:
                    continue
                
                # Extract work information - prefer code column if available
                work_code = None
                work_name = None
                
                # Check the code column (usually column before the name column)
                if work_name_col > 0:
                    code_col_value = self._get_cell_value(ws, row_idx, work_name_col)  # work_name_col is 0-based, so this checks the previous column
                    if code_col_value and isinstance(code_col_value, str):
                        code_col_value = code_col_value.strip()
                        # Check if this looks like a work code (contains dots and numbers)
                        if '.' in code_col_value and any(c.isdigit() for c in code_col_value):
                            work_code = code_col_value
                
                # Get the work name from the detected column
                name_col_value = self._get_cell_value(ws, row_idx, work_name_col + 1)
                if name_col_value and isinstance(name_col_value, str):
                    work_name = name_col_value.strip()
                
                # Use code if available, otherwise use name
                work_info = work_code if work_code else work_name
                if not work_info:
                    continue
                
                # Extract quantity (actual labor)
                try:
                    quantity_value = self._get_cell_value(ws, row_idx, quantity_col + 1) if quantity_col is not None else None
                    actual_labor = float(quantity_value) if quantity_value else 0.0
                except (ValueError, TypeError):
                    actual_labor = 0.0
                
                # Skip rows with zero quantity
                if actual_labor == 0.0:
                    continue
                
                line_number += 1
                
                # Try to find work by code or name
                work_id = self._find_work_by_code_or_name(work_info)
                planned_labor = 0.0
                
                if work_id:
                    # Get planned labor from estimate if available
                    if daily_report.estimate_id:
                        planned_labor = self._get_planned_labor_from_estimate(daily_report.estimate_id, work_id)
                
                # Calculate deviation if we have planned labor
                deviation_percent = 0.0
                if planned_labor and planned_labor > 0:
                    deviation_percent = ((actual_labor - planned_labor) / planned_labor) * 100
                
                # Create dataclass DailyReportLine
                line = DailyReportLine(
                    line_number=line_number,
                    work_id=work_id or 0,
                    planned_labor=planned_labor,
                    actual_labor=actual_labor,
                    deviation_percent=deviation_percent,
                    executor_ids=[]  # Empty list for now, can be filled later
                )
                
                daily_report.lines.append(line)
            
            if not daily_report.lines:
                return None, "Не найдено ни одной строки с работами"
            
            return daily_report, ""
            
        except Exception as e:
            return None, f"Ошибка при импорте: {str(e)}"
    
    def _extract_estimate_reference(self, text: str) -> Optional[str]:
        """Extract estimate reference from text"""
        text = text.strip()
        
        # Look for patterns like "Смета №123" or "№123"
        import re
        patterns = [
            r'№\s*([^\s,]+)',
            r'смета\s*№?\s*([^\s,]+)',
            r'локальная\s+смета\s*№?\s*([^\s,]+)'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                return match.group(1).strip()
        
        return None
    
    def _extract_foreman_name(self, text: str) -> Optional[str]:
        """Extract foreman name from text"""
        text = text.strip()
        
        # Look for patterns like "Бригадир: Иванов И.И."
        import re
        patterns = [
            r'бригадир:?\s*(.+)',
            r'прораб:?\s*(.+)'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                return match.group(1).strip()
        
        return None
    
    def _extract_date(self, text: str) -> Optional[date]:
        """Extract date from text"""
        import re
        from datetime import datetime
        
        # Look for date patterns
        patterns = [
            r'(\d{1,2})[./](\d{1,2})[./](\d{4})',
            r'(\d{1,2})[./](\d{1,2})[./](\d{2})'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                try:
                    day, month, year = match.groups()
                    year = int(year)
                    if year < 100:
                        year += 2000
                    return date(year, int(month), int(day))
                except ValueError:
                    continue
        
        return None
    
    def _find_estimate_by_number(self, number: str) -> Optional[int]:
        """Find estimate by number"""
        cursor = self.db.cursor()
        cursor.execute("SELECT id FROM estimates WHERE number = ? AND marked_for_deletion = 0", (number,))
        row = cursor.fetchone()
        return row['id'] if row else None
    
    def _find_or_create_person(self, name: str) -> Optional[int]:
        """Find or create person"""
        if not name:
            return None
        
        cursor = self.db.cursor()
        
        # Try to find existing
        cursor.execute("SELECT id FROM persons WHERE full_name = ? AND marked_for_deletion = 0", (name,))
        row = cursor.fetchone()
        if row:
            return row['id']
        
        # Create new
        cursor.execute("""
            INSERT INTO persons (full_name, marked_for_deletion)
            VALUES (?, 0)
        """, (name,))
        self.db.commit()
        return cursor.lastrowid
    
    def _find_work_by_code_or_name(self, work_info: str) -> Optional[int]:
        """Find work by code or name"""
        if not work_info:
            return None
        
        cursor = self.db.cursor()
        
        # First try to extract code (usually at the beginning)
        import re
        code_match = re.match(r'^([^\s]+)', work_info.strip())
        if code_match:
            code = code_match.group(1)
            cursor.execute("SELECT id FROM works WHERE code = ? AND marked_for_deletion = 0", (code,))
            row = cursor.fetchone()
            if row:
                return row['id']
        
        # Try to find by name (partial match)
        cursor.execute("""
            SELECT id FROM works 
            WHERE name LIKE ? AND marked_for_deletion = 0
            ORDER BY LENGTH(name)
            LIMIT 1
        """, (f"%{work_info}%",))
        row = cursor.fetchone()
        return row['id'] if row else None
    
    def _get_planned_labor_from_estimate(self, estimate_id: int, work_id: int) -> float:
        """Get planned labor from estimate for specific work"""
        cursor = self.db.cursor()
        cursor.execute("""
            SELECT planned_labor FROM estimate_lines 
            WHERE estimate_id = ? AND work_id = ?
        """, (estimate_id, work_id))
        row = cursor.fetchone()
        return row['planned_labor'] if row else 0.0
    
    def _get_cell_value(self, ws, row_idx: int, col_idx: int):
        """Get cell value, handling merged cells"""
        cell = ws.cell(row=row_idx, column=col_idx)
        
        # Check if cell is part of a merged range
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # Get the top-left cell of the merged range
                top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                return top_left_cell.value
        
        return cell.value