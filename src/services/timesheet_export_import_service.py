"""Service for exporting and importing timesheet data"""
import json
import os
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from src.data.database_manager import DatabaseManager
from src.data.models.sqlalchemy_models import Timesheet, TimesheetLine, Person, Object, Estimate


class TimesheetExportImportService:
    """Service for timesheet export/import operations"""
    
    def __init__(self):
        self.db_manager = DatabaseManager()
    
    def export_timesheet_to_json(self, timesheet_id: int) -> Optional[Dict[str, Any]]:
        """Export timesheet data to JSON format"""
        try:
            with self.db_manager.get_session() as session:
                timesheet = session.query(Timesheet).filter_by(id=timesheet_id).first()
                if not timesheet:
                    return None
                
                # Get related data
                object_name = timesheet.object.name if timesheet.object else None
                estimate_number = timesheet.estimate.number if timesheet.estimate else None
                foreman_name = timesheet.foreman.full_name if timesheet.foreman else None
                
                # Export main document data
                timesheet_data = {
                    'document_info': {
                        'type': 'timesheet',
                        'version': '1.0',
                        'exported_at': datetime.now().isoformat(),
                        'number': timesheet.number,
                        'date': timesheet.date.isoformat(),
                        'month_year': timesheet.month_year,
                        'is_posted': timesheet.is_posted,
                        'object_name': object_name,
                        'estimate_number': estimate_number,
                        'foreman_name': foreman_name
                    },
                    'lines': []
                }
                
                # Export lines
                for line in timesheet.lines:
                    employee_name = line.employee.full_name if line.employee else None
                    
                    line_data = {
                        'line_number': line.line_number,
                        'employee_name': employee_name,
                        'hourly_rate': float(line.hourly_rate) if line.hourly_rate else 0.0,
                        'days_data': {}
                    }
                    
                    # Export daily hours
                    for i in range(1, 32):  # Days 1-31
                        day_field = f'day_{i:02d}'
                        hours = getattr(line, day_field, None)
                        if hours is not None:
                            line_data['days_data'][day_field] = float(hours)
                    
                    line_data['total_hours'] = float(line.total_hours) if line.total_hours else 0.0
                    line_data['total_amount'] = float(line.total_amount) if line.total_amount else 0.0
                    
                    timesheet_data['lines'].append(line_data)
                
                return timesheet_data
                
        except Exception as e:
            print(f"Error exporting timesheet to JSON: {e}")
            return None
    
    def export_timesheet_to_excel(self, timesheet_id: int, file_path: str) -> bool:
        """Export timesheet to Excel format with print form"""
        try:
            timesheet_data = self.export_timesheet_to_json(timesheet_id)
            if not timesheet_data:
                return False
            
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Табель"
            
            # Styles
            title_font = Font(name='Arial', size=14, bold=True)
            header_font = Font(name='Arial', size=10, bold=True)
            data_font = Font(name='Arial', size=9)
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
            
            current_row = 1
            doc_info = timesheet_data['document_info']
            
            # Title
            sheet.merge_cells(f'A{current_row}:AH{current_row}')
            title_cell = sheet[f'A{current_row}']
            title_cell.value = f"ТАБЕЛЬ УЧЕТА РАБОЧЕГО ВРЕМЕНИ №{doc_info['number']}"
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 2
            
            # Document info
            info_data = [
                ('Дата документа:', doc_info['date']),
                ('Период:', doc_info['month_year']),
                ('Объект:', doc_info['object_name'] or ''),
                ('Смета:', doc_info['estimate_number'] or ''),
                ('Прораб:', doc_info['foreman_name'] or ''),
                ('Проведен:', 'Да' if doc_info['is_posted'] else 'Нет')
            ]
            
            for label, value in info_data:
                sheet[f'A{current_row}'] = label
                sheet[f'B{current_row}'] = value
                sheet[f'A{current_row}'].font = header_font
                current_row += 1
            
            current_row += 1
            
            # Table headers
            headers = ['№', 'Сотрудник', 'Ставка']
            
            # Add day columns (1-31)
            for day in range(1, 32):
                headers.append(str(day))
            
            headers.extend(['Итого часов', 'Сумма'])
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=current_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1
            
            # Write data rows
            for line in timesheet_data['lines']:
                col = 1
                
                # Line number
                cell = sheet.cell(row=current_row, column=col)
                cell.value = line['line_number']
                cell.font = data_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                col += 1
                
                # Employee name
                cell = sheet.cell(row=current_row, column=col)
                cell.value = line['employee_name'] or ''
                cell.font = data_font
                cell.border = border
                col += 1
                
                # Hourly rate
                cell = sheet.cell(row=current_row, column=col)
                cell.value = line['hourly_rate']
                cell.font = data_font
                cell.border = border
                cell.alignment = Alignment(horizontal='right')
                col += 1
                
                # Daily hours (days 1-31)
                for day in range(1, 32):
                    day_field = f'day_{day:02d}'
                    hours = line['days_data'].get(day_field, '')
                    
                    cell = sheet.cell(row=current_row, column=col)
                    cell.value = hours if hours else ''
                    cell.font = data_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                    col += 1
                
                # Total hours
                cell = sheet.cell(row=current_row, column=col)
                cell.value = line['total_hours']
                cell.font = data_font
                cell.border = border
                cell.alignment = Alignment(horizontal='right')
                col += 1
                
                # Total amount
                cell = sheet.cell(row=current_row, column=col)
                cell.value = line['total_amount']
                cell.font = data_font
                cell.border = border
                cell.alignment = Alignment(horizontal='right')
                
                current_row += 1
            
            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                if col == 2:  # Employee name column
                    sheet.column_dimensions[get_column_letter(col)].width = 25
                elif col == 3:  # Rate column
                    sheet.column_dimensions[get_column_letter(col)].width = 10
                elif col > 3 and col <= 34:  # Day columns
                    sheet.column_dimensions[get_column_letter(col)].width = 4
                else:  # Other columns
                    sheet.column_dimensions[get_column_letter(col)].width = 12
            
            workbook.save(file_path)
            return True
            
        except Exception as e:
            print(f"Error exporting timesheet to Excel: {e}")
            return False
    
    def import_timesheet_from_excel(self, file_path: str, timesheet_id: int) -> Tuple[bool, str]:
        """Import timesheet data from Excel file"""
        try:
            if not os.path.exists(file_path):
                return False, "Файл не найден"
            
            workbook = load_workbook(file_path)
            sheet = workbook.active
            
            with self.db_manager.get_session() as session:
                timesheet = session.query(Timesheet).filter_by(id=timesheet_id).first()
                if not timesheet:
                    return False, "Табель не найден"
                
                if timesheet.is_posted:
                    return False, "Нельзя импортировать данные в проведенный документ"
                
                # Find data start row (look for row with "№" in first column)
                data_start_row = None
                for row in range(1, 20):  # Search in first 20 rows
                    cell_value = sheet.cell(row=row, column=1).value
                    if cell_value == "№":
                        data_start_row = row + 1
                        break
                
                if not data_start_row:
                    return False, "Не найден заголовок таблицы данных"
                
                # Clear existing lines
                for line in timesheet.lines:
                    session.delete(line)
                session.flush()
                
                # Import data rows
                line_number = 1
                row = data_start_row
                
                while True:
                    # Check if we have employee name
                    employee_name = sheet.cell(row=row, column=2).value
                    if not employee_name:
                        break
                    
                    # Find employee by name
                    employee = session.query(Person).filter_by(full_name=employee_name.strip()).first()
                    if not employee:
                        return False, f"Сотрудник '{employee_name}' не найден в базе данных"
                    
                    # Get hourly rate
                    hourly_rate = sheet.cell(row=row, column=3).value or 0
                    
                    # Create new line
                    line = TimesheetLine(
                        timesheet_id=timesheet_id,
                        line_number=line_number,
                        employee_id=employee.id,
                        hourly_rate=float(hourly_rate)
                    )
                    
                    # Import daily hours
                    total_hours = 0
                    for day in range(1, 32):
                        col = 3 + day  # Days start from column 4
                        hours_value = sheet.cell(row=row, column=col).value
                        
                        if hours_value:
                            try:
                                hours = float(hours_value)
                                day_field = f'day_{day:02d}'
                                setattr(line, day_field, hours)
                                total_hours += hours
                            except (ValueError, TypeError):
                                pass
                    
                    line.total_hours = total_hours
                    line.total_amount = total_hours * float(hourly_rate)
                    
                    session.add(line)
                    line_number += 1
                    row += 1
                
                session.commit()
                return True, f"Успешно импортировано {line_number - 1} строк"
                
        except Exception as e:
            return False, f"Ошибка импорта: {str(e)}"
    
    def export_timesheet_to_json_file(self, timesheet_id: int, file_path: str) -> bool:
        """Export timesheet to JSON file"""
        try:
            data = self.export_timesheet_to_json(timesheet_id)
            if not data:
                return False
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            return True
            
        except Exception as e:
            print(f"Error exporting timesheet to JSON file: {e}")
            return False
    
    def import_timesheet_from_json_file(self, file_path: str, timesheet_id: int) -> Tuple[bool, str]:
        """Import timesheet from JSON file"""
        try:
            if not os.path.exists(file_path):
                return False, "Файл не найден"
            
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            if data.get('document_info', {}).get('type') != 'timesheet':
                return False, "Неверный формат файла"
            
            with self.db_manager.get_session() as session:
                timesheet = session.query(Timesheet).filter_by(id=timesheet_id).first()
                if not timesheet:
                    return False, "Табель не найден"
                
                if timesheet.is_posted:
                    return False, "Нельзя импортировать данные в проведенный документ"
                
                # Clear existing lines
                for line in timesheet.lines:
                    session.delete(line)
                session.flush()
                
                # Import lines
                imported_count = 0
                for line_data in data.get('lines', []):
                    employee_name = line_data.get('employee_name')
                    if not employee_name:
                        continue
                    
                    # Find employee by name
                    employee = session.query(Person).filter_by(full_name=employee_name.strip()).first()
                    if not employee:
                        continue
                    
                    # Create new line
                    line = TimesheetLine(
                        timesheet_id=timesheet_id,
                        line_number=line_data.get('line_number', imported_count + 1),
                        employee_id=employee.id,
                        hourly_rate=line_data.get('hourly_rate', 0)
                    )
                    
                    # Import daily hours
                    for day_field, hours in line_data.get('days_data', {}).items():
                        if hasattr(line, day_field):
                            setattr(line, day_field, hours)
                    
                    line.total_hours = line_data.get('total_hours', 0)
                    line.total_amount = line_data.get('total_amount', 0)
                    
                    session.add(line)
                    imported_count += 1
                
                session.commit()
                return True, f"Успешно импортировано {imported_count} строк"
                
        except Exception as e:
            return False, f"Ошибка импорта: {str(e)}"