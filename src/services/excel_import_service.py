"""Excel import service"""
import openpyxl
from datetime import date
from typing import Optional, Tuple
from ..data.models.estimate import Estimate, EstimateLine
from ..data.database_manager import DatabaseManager


class ExcelImportService:
    def __init__(self):
        self.db = DatabaseManager().get_connection()
    
    def import_estimate(self, file_path: str) -> Tuple[Optional[Estimate], str]:
        """
        Import estimate from Excel file
        Returns: (Estimate object or None, error message)
        """
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            estimate = Estimate()
            estimate.date = date.today()
            
            # Parse header information
            for row in ws.iter_rows(min_row=1, max_row=15):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        # Extract customer
                        if "Заказчик:" in cell.value:
                            customer_name = cell.value.replace("Заказчик:", "").strip()
                            customer_id = self._find_or_create_counterparty(customer_name)
                            if customer_id:
                                estimate.customer_id = customer_id
                        
                        # Extract contractor
                        elif "Подрядчик:" in cell.value:
                            contractor_name = cell.value.replace("Подрядчик:", "").strip()
                            contractor_id = self._find_or_create_organization(contractor_name)
                            if contractor_id:
                                estimate.contractor_id = contractor_id
                        
                        # Extract estimate number
                        elif "ЛОКАЛЬНЫЙ СМЕТНЫЙ РАСЧЕТ" in cell.value or "локальная смета" in cell.value.lower():
                            # Try to extract number
                            parts = cell.value.split("№")
                            if len(parts) > 1:
                                estimate.number = parts[1].strip()
                        
                        # Extract object name
                        elif len(cell.value) > 50 and "наименование" not in cell.value.lower():
                            # This might be the object description
                            if not estimate.number:  # Use as object name if we haven't found it yet
                                object_name = cell.value.strip()
                                if estimate.customer_id:
                                    object_id = self._find_or_create_object(object_name, estimate.customer_id)
                                    if object_id:
                                        estimate.object_id = object_id
            
            # If no number found, generate one
            if not estimate.number:
                estimate.number = f"ИМП-{date.today().strftime('%Y%m%d')}"
            
            # Find the header row (contains "Наименование работ")
            header_row = None
            for row_idx, row in enumerate(ws.iter_rows(min_row=15, max_row=25), start=15):
                for cell in row:
                    if cell.value and "Наименование работ" in str(cell.value):
                        header_row = row_idx
                        break
                if header_row:
                    break
            
            if not header_row:
                return None, "Не найден заголовок таблицы с работами"
            
            # Parse table rows
            line_number = 0
            for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
                # Check if this is a data row (first cell should be a number)
                if not row[0].value:
                    continue
                
                try:
                    # Try to parse as line number
                    int(str(row[0].value))
                except (ValueError, TypeError):
                    continue
                
                line = EstimateLine()
                line_number += 1
                line.line_number = line_number
                
                # Extract data from columns
                # Column structure: N, Код, Наименование, ЕдИзм, Кол-во, ..., Всего, ТЗ
                work_code = str(row[1].value).strip() if row[1].value else ""
                work_name = str(row[2].value) if row[2].value else ""
                unit = str(row[3].value).strip() if row[3].value else ""
                
                try:
                    quantity = float(row[4].value) if row[4].value else 0.0
                except (ValueError, TypeError):
                    quantity = 0.0
                
                # Find price (Всего на единицу работ - column 7)
                try:
                    price = float(row[7].value) if row[7].value else 0.0
                except (ValueError, TypeError):
                    price = 0.0
                
                # Find labor rate (ТЗ на единицу работ - column 8)
                try:
                    labor_rate = float(row[8].value) if row[8].value else 0.0
                except (ValueError, TypeError):
                    labor_rate = 0.0
                
                # Find or create work
                work_id = self._find_or_create_work(work_name, work_code, unit, price, labor_rate)
                if not work_id:
                    continue
                
                line.work_id = work_id
                line.quantity = quantity
                line.unit = unit  # This is EstimateLine.unit, not Work.unit
                line.price = price
                line.labor_rate = labor_rate
                line.sum = quantity * price
                line.planned_labor = quantity * labor_rate
                
                estimate.lines.append(line)
            
            # Calculate totals
            estimate.total_sum = sum(line.sum for line in estimate.lines)
            estimate.total_labor = sum(line.planned_labor for line in estimate.lines)
            
            if not estimate.lines:
                return None, "Не найдено ни одной строки с работами"
            
            return estimate, ""
            
        except Exception as e:
            return None, f"Ошибка при импорте: {str(e)}"
    
    def _find_or_create_counterparty(self, name: str) -> Optional[int]:
        """Find or create counterparty"""
        if not name:
            return None
        
        cursor = self.db.cursor()
        
        # Try to find existing
        cursor.execute("SELECT id FROM counterparties WHERE name = ?", (name,))
        row = cursor.fetchone()
        if row:
            return row['id']
        
        # Create new
        cursor.execute("""
            INSERT INTO counterparties (name, marked_for_deletion)
            VALUES (?, 0)
        """, (name,))
        self.db.commit()
        return cursor.lastrowid
    
    def _find_or_create_organization(self, name: str) -> Optional[int]:
        """Find or create organization"""
        if not name:
            return None
        
        cursor = self.db.cursor()
        
        # Try to find existing
        cursor.execute("SELECT id FROM organizations WHERE name = ?", (name,))
        row = cursor.fetchone()
        if row:
            return row['id']
        
        # Create new
        cursor.execute("""
            INSERT INTO organizations (name, marked_for_deletion)
            VALUES (?, 0)
        """, (name,))
        self.db.commit()
        return cursor.lastrowid
    
    def _find_or_create_object(self, name: str, owner_id: int) -> Optional[int]:
        """Find or create object"""
        if not name:
            return None
        
        cursor = self.db.cursor()
        
        # Try to find existing
        cursor.execute("SELECT id FROM objects WHERE name = ? AND owner_id = ?", (name, owner_id))
        row = cursor.fetchone()
        if row:
            return row['id']
        
        # Create new
        cursor.execute("""
            INSERT INTO objects (name, owner_id, marked_for_deletion)
            VALUES (?, ?, 0)
        """, (name, owner_id))
        self.db.commit()
        return cursor.lastrowid
    
    def _find_or_create_work(self, name: str, code: str, unit: str, price: float, labor_rate: float) -> Optional[int]:
        """Find or create work"""
        if not name:
            return None
        
        cursor = self.db.cursor()
        
        # Try to find existing by code if provided
        if code:
            cursor.execute("SELECT id FROM works WHERE code = ?", (code,))
            row = cursor.fetchone()
            if row:
                # Update existing work with new data
                cursor.execute("""
                    UPDATE works 
                    SET name = ?, unit = ?, price = ?, labor_rate = ?
                    WHERE id = ?
                """, (name, unit, price, labor_rate, row['id']))
                self.db.commit()
                return row['id']
        
        # Try to find existing by name
        cursor.execute("SELECT id FROM works WHERE name = ?", (name,))
        row = cursor.fetchone()
        if row:
            # Update code if provided
            if code:
                cursor.execute("UPDATE works SET code = ? WHERE id = ?", (code, row['id']))
                self.db.commit()
            return row['id']
        
        # Create new
        cursor.execute("""
            INSERT INTO works (name, code, unit, price, labor_rate, marked_for_deletion)
            VALUES (?, ?, ?, ?, ?, 0)
        """, (name, code, unit, price, labor_rate))
        self.db.commit()
        return cursor.lastrowid
