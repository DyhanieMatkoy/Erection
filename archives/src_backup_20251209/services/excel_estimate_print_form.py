"""Excel estimate print form generator"""
from typing import Optional
from openpyxl.styles import Font, Alignment, Border, Side
from .excel_print_form_generator import ExcelPrintFormGenerator
from ..data.database_manager import DatabaseManager


class ExcelEstimatePrintForm(ExcelPrintFormGenerator):
    """Generator for estimate print forms in Excel format"""
    
    TEMPLATE_NAME = "estimate_template.xlsx"
    
    def __init__(self):
        """Initialize Excel estimate print form generator"""
        super().__init__()
        self.db = DatabaseManager().get_connection()
    
    def generate(self, estimate_id: int) -> Optional[bytes]:
        """
        Generate estimate print form in Excel format
        
        Args:
            estimate_id: ID of the estimate
            
        Returns:
            Excel content as bytes or None if estimate not found
        """
        # Load estimate data
        estimate_data = self._load_estimate_data(estimate_id)
        if not estimate_data:
            return None
        
        # Try to load template, otherwise create from scratch
        if self.template_exists(self.TEMPLATE_NAME):
            workbook = self.load_template(self.TEMPLATE_NAME)
            if workbook:
                self._fill_template(workbook, estimate_data)
            else:
                workbook = self._create_from_scratch(estimate_data)
        else:
            workbook = self._create_from_scratch(estimate_data)
        
        return self.save_to_bytes(workbook)
    
    def _load_estimate_data(self, estimate_id: int) -> Optional[dict]:
        """Load estimate data from database"""
        cursor = self.db.cursor()
        
        # Load estimate header
        cursor.execute("""
            SELECT 
                e.id, e.number, e.date, e.total_sum, e.total_labor,
                c.name as customer_name,
                o.name as object_name,
                org.name as contractor_name,
                p.full_name as responsible_name
            FROM estimates e
            LEFT JOIN counterparties c ON e.customer_id = c.id
            LEFT JOIN objects o ON e.object_id = o.id
            LEFT JOIN organizations org ON e.contractor_id = org.id
            LEFT JOIN persons p ON e.responsible_id = p.id
            WHERE e.id = ?
        """, (estimate_id,))
        
        row = cursor.fetchone()
        if not row:
            return None
        
        estimate_data = {
            'id': row['id'],
            'number': row['number'],
            'date': row['date'],
            'total_sum': row['total_sum'],
            'total_labor': row['total_labor'],
            'customer_name': row['customer_name'],
            'object_name': row['object_name'],
            'contractor_name': row['contractor_name'],
            'responsible_name': row['responsible_name'],
            'lines': []
        }
        
        # Load estimate lines
        cursor.execute("""
            SELECT 
                el.line_number,
                el.work_id,
                w.name as work_name,
                w.code as work_code,
                el.quantity,
                el.unit,
                el.price,
                el.labor_rate,
                el.sum,
                el.planned_labor
            FROM estimate_lines el
            LEFT JOIN works w ON el.work_id = w.id
            WHERE el.estimate_id = ?
            ORDER BY el.line_number
        """, (estimate_id,))
        
        for line_row in cursor.fetchall():
            is_group = line_row['work_id'] == -1
            
            estimate_data['lines'].append({
                'line_number': line_row['line_number'],
                'work_name': line_row['work_name'] or line_row['unit'] if not is_group else "",
                'work_code': line_row['work_code'] or "",
                'quantity': line_row['quantity'],
                'unit': line_row['unit'] if not is_group else "",
                'price': line_row['price'],
                'labor_rate': line_row['labor_rate'],
                'sum': line_row['sum'],
                'planned_labor': line_row['planned_labor'],
                'is_group': is_group,
                'group_name': line_row['unit'] if is_group else ""
            })
        
        return estimate_data
    
    def _fill_template(self, workbook, estimate_data: dict):
        """Fill template with data"""
        sheet = workbook.active
        
        # Try to find and fill placeholders
        # This is a simple implementation - can be enhanced based on template structure
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    value = str(cell.value)
                    # Replace placeholders
                    value = value.replace('{NUMBER}', estimate_data['number'] or '')
                    value = value.replace('{DATE}', self.format_date(estimate_data['date']))
                    value = value.replace('{OBJECT}', estimate_data['object_name'] or '')
                    value = value.replace('{CUSTOMER}', estimate_data['customer_name'] or '')
                    value = value.replace('{CONTRACTOR}', estimate_data['contractor_name'] or '')
                    value = value.replace('{TOTAL_SUM}', self.format_number(estimate_data['total_sum']))
                    value = value.replace('{TOTAL_LABOR}', self.format_number(estimate_data['total_labor']))
                    cell.value = value
        
        # Find table start (look for header row with "N" or "№")
        table_start_row = None
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=50), start=1):
            for cell in row:
                if cell.value and str(cell.value).strip() in ['N', '№', 'N п/п', '№ п/п']:
                    table_start_row = row_idx + 1
                    break
            if table_start_row:
                break
        
        # Fill lines if table found
        if table_start_row:
            self._fill_lines_in_template(sheet, estimate_data['lines'], table_start_row)
    
    def _fill_lines_in_template(self, sheet, lines: list, start_row: int):
        """Fill estimate lines in template"""
        current_row = start_row
        
        for line in lines:
            if line.get('is_group', False):
                # Group row
                self.set_cell_value(sheet, current_row, 1, line['line_number'])
                self.set_cell_value(sheet, current_row, 2, line.get('group_name', ''))
                # Make group row bold
                for col in range(1, 14):
                    cell = sheet.cell(row=current_row, column=col)
                    cell.font = Font(bold=True, size=10)
            else:
                # Regular line
                unit_labor = line['labor_rate'] if line['quantity'] > 0 else 0
                unit_materials = (line['price'] - unit_labor) if line['price'] > unit_labor else 0
                unit_total = line['price']
                unit_tz = line['labor_rate'] / line['quantity'] if line['quantity'] > 0 else 0
                
                total_labor = unit_labor * line['quantity']
                total_materials = unit_materials * line['quantity']
                total_sum = line['sum']
                total_tz = line['planned_labor']
                
                self.set_cell_value(sheet, current_row, 1, line['line_number'])
                self.set_cell_value(sheet, current_row, 2, line.get('work_code', ''))
                self.set_cell_value(sheet, current_row, 3, line['work_name'])
                self.set_cell_value(sheet, current_row, 4, line['unit'])
                self.set_cell_value(sheet, current_row, 5, float(line['quantity']))
                self.set_cell_value(sheet, current_row, 6, float(unit_labor))
                self.set_cell_value(sheet, current_row, 7, float(unit_materials))
                self.set_cell_value(sheet, current_row, 8, float(unit_total))
                self.set_cell_value(sheet, current_row, 9, float(unit_tz))
                self.set_cell_value(sheet, current_row, 10, float(total_labor))
                self.set_cell_value(sheet, current_row, 11, float(total_materials))
                self.set_cell_value(sheet, current_row, 12, float(total_sum))
                self.set_cell_value(sheet, current_row, 13, float(total_tz))
            
            current_row += 1
    
    def _create_from_scratch(self, estimate_data: dict):
        """Create Excel document from scratch"""
        workbook = self.create_workbook()
        sheet = workbook.active
        sheet.title = "Смета"
        
        current_row = 1
        
        # Approval section
        self.set_cell_value(sheet, current_row, 1, "Согласовано:")
        self.set_cell_value(sheet, current_row, 6, "Утверждаю:")
        current_row += 1
        
        customer = estimate_data['customer_name'] or "_______________"
        contractor = estimate_data['contractor_name'] or "_______________"
        self.set_cell_value(sheet, current_row, 1, f"Заказчик: {customer}")
        self.set_cell_value(sheet, current_row, 6, f"Подрядчик: {contractor}")
        current_row += 3
        
        # Title
        title_row = current_row
        self.set_cell_value(sheet, title_row, 1, f"ЛОКАЛЬНЫЙ СМЕТНЫЙ РАСЧЕТ №{estimate_data['number']}")
        self.merge_cells(sheet, title_row, 1, title_row, 13)
        title_style = self.create_title_style()
        self.set_cell_style(sheet, title_row, 1, **title_style)
        current_row += 1
        
        self.set_cell_value(sheet, current_row, 1, "(локальная смета)")
        self.merge_cells(sheet, current_row, 1, current_row, 13)
        self.set_cell_style(sheet, current_row, 1, 
                           alignment=Alignment(horizontal='center', vertical='center'))
        current_row += 2
        
        # Object
        object_desc = estimate_data['object_name'] or "Объект не указан"
        self.set_cell_value(sheet, current_row, 1, object_desc)
        self.merge_cells(sheet, current_row, 1, current_row, 13)
        self.set_cell_style(sheet, current_row, 1,
                           alignment=Alignment(horizontal='center', vertical='center'))
        current_row += 1
        
        self.set_cell_value(sheet, current_row, 1, "(наименование работ и затрат, наименование объекта)")
        self.merge_cells(sheet, current_row, 1, current_row, 13)
        self.set_cell_style(sheet, current_row, 1,
                           font=Font(size=9, italic=True),
                           alignment=Alignment(horizontal='center', vertical='center'))
        current_row += 2
        
        # Table header
        header_row1 = current_row
        headers1 = ["", "", "", "", "", "на единицу работ", "", "", "", "на полный объём работ", "", "", ""]
        for col, header in enumerate(headers1, start=1):
            self.set_cell_value(sheet, header_row1, col, header)
            header_style = self.create_header_style()
            self.set_cell_style(sheet, header_row1, col, **header_style)
        
        self.merge_cells(sheet, header_row1, 6, header_row1, 9)
        self.merge_cells(sheet, header_row1, 10, header_row1, 13)
        current_row += 1
        
        header_row2 = current_row
        headers2 = ["N", "Код", "Наименование работ и затрат", "ЕдИзм", "Кол-во",
                   "Зарплата", "Материалы,\nмеханизмы", "Всего", "ТЗ",
                   "Зарплата", "Материалы,\nмеханизмы", "Всего", "ТЗ"]
        for col, header in enumerate(headers2, start=1):
            self.set_cell_value(sheet, header_row2, col, header)
            header_style = self.create_header_style()
            self.set_cell_style(sheet, header_row2, col, **header_style)
        current_row += 1
        
        # Set column widths
        widths = [5, 10, 35, 8, 8, 10, 10, 10, 8, 10, 10, 10, 8]
        for col, width in enumerate(widths, start=1):
            self.set_column_width(sheet, col, width)
        
        # Data rows
        data_style = self.create_data_style()
        number_style = self.create_number_style()
        
        for line in estimate_data['lines']:
            if line.get('is_group', False):
                # Group row
                self.set_cell_value(sheet, current_row, 1, line['line_number'])
                self.set_cell_value(sheet, current_row, 2, line.get('group_name', ''))
                self.merge_cells(sheet, current_row, 2, current_row, 13)
                for col in range(1, 14):
                    cell = sheet.cell(row=current_row, column=col)
                    cell.font = Font(bold=True, size=10)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            else:
                # Regular line
                unit_labor = line['labor_rate'] if line['quantity'] > 0 else 0
                unit_materials = (line['price'] - unit_labor) if line['price'] > unit_labor else 0
                unit_total = line['price']
                unit_tz = line['labor_rate'] / line['quantity'] if line['quantity'] > 0 else 0
                
                total_labor = unit_labor * line['quantity']
                total_materials = unit_materials * line['quantity']
                total_sum = line['sum']
                total_tz = line['planned_labor']
                
                self.set_cell_value(sheet, current_row, 1, line['line_number'])
                self.set_cell_value(sheet, current_row, 2, line.get('work_code', ''))
                self.set_cell_value(sheet, current_row, 3, line['work_name'])
                self.set_cell_value(sheet, current_row, 4, line['unit'])
                self.set_cell_value(sheet, current_row, 5, float(line['quantity']))
                self.set_cell_value(sheet, current_row, 6, float(unit_labor))
                self.set_cell_value(sheet, current_row, 7, float(unit_materials))
                self.set_cell_value(sheet, current_row, 8, float(unit_total))
                self.set_cell_value(sheet, current_row, 9, float(unit_tz))
                self.set_cell_value(sheet, current_row, 10, float(total_labor))
                self.set_cell_value(sheet, current_row, 11, float(total_materials))
                self.set_cell_value(sheet, current_row, 12, float(total_sum))
                self.set_cell_value(sheet, current_row, 13, float(total_tz))
                
                # Apply styles
                for col in range(1, 14):
                    if col in [1, 2, 3, 4]:
                        self.set_cell_style(sheet, current_row, col, **data_style)
                    else:
                        self.set_cell_style(sheet, current_row, col, **number_style)
            
            current_row += 1
        
        # Totals
        current_row += 1
        self.set_cell_value(sheet, current_row, 1, "ИТОГО по смете:")
        self.set_cell_style(sheet, current_row, 1, font=Font(bold=True, size=11))
        current_row += 1
        
        self.set_cell_value(sheet, current_row, 1, 
                           f"Всего: {self.format_number(estimate_data['total_sum'])} руб.")
        current_row += 1
        
        self.set_cell_value(sheet, current_row, 1,
                           f"Трудозатраты: {self.format_number(estimate_data['total_labor'])} ч.")
        
        return workbook
    
    def create_template(self) -> bool:
        """Create default template file"""
        try:
            # Create a sample template with placeholders
            workbook = self.create_workbook()
            sheet = workbook.active
            sheet.title = "Смета"
            
            # Add placeholders and structure
            self.set_cell_value(sheet, 1, 1, "Согласовано:")
            self.set_cell_value(sheet, 1, 6, "Утверждаю:")
            self.set_cell_value(sheet, 2, 1, "Заказчик: {CUSTOMER}")
            self.set_cell_value(sheet, 2, 6, "Подрядчик: {CONTRACTOR}")
            
            self.set_cell_value(sheet, 5, 1, "ЛОКАЛЬНЫЙ СМЕТНЫЙ РАСЧЕТ №{NUMBER}")
            self.merge_cells(sheet, 5, 1, 5, 13)
            title_style = self.create_title_style()
            self.set_cell_style(sheet, 5, 1, **title_style)
            
            self.set_cell_value(sheet, 6, 1, "(локальная смета)")
            self.merge_cells(sheet, 6, 1, 6, 13)
            
            self.set_cell_value(sheet, 8, 1, "{OBJECT}")
            self.merge_cells(sheet, 8, 1, 8, 13)
            
            # Table header
            header_row1 = 11
            headers1 = ["", "", "", "", "", "на единицу работ", "", "", "", "на полный объём работ", "", "", ""]
            for col, header in enumerate(headers1, start=1):
                self.set_cell_value(sheet, header_row1, col, header)
                header_style = self.create_header_style()
                self.set_cell_style(sheet, header_row1, col, **header_style)
            
            self.merge_cells(sheet, header_row1, 6, header_row1, 9)
            self.merge_cells(sheet, header_row1, 10, header_row1, 13)
            
            header_row2 = 12
            headers2 = ["N", "Код", "Наименование работ и затрат", "ЕдИзм", "Кол-во",
                       "Зарплата", "Материалы,\nмеханизмы", "Всего", "ТЗ",
                       "Зарплата", "Материалы,\nмеханизмы", "Всего", "ТЗ"]
            for col, header in enumerate(headers2, start=1):
                self.set_cell_value(sheet, header_row2, col, header)
                header_style = self.create_header_style()
                self.set_cell_style(sheet, header_row2, col, **header_style)
            
            # Set column widths
            widths = [5, 10, 35, 8, 8, 10, 10, 10, 8, 10, 10, 10, 8]
            for col, width in enumerate(widths, start=1):
                self.set_column_width(sheet, col, width)
            
            # Save template
            template_path = self.get_template_path(self.TEMPLATE_NAME)
            workbook.save(template_path)
            return True
        except Exception as e:
            print(f"Error creating template: {e}")
            return False
