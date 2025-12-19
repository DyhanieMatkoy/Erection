"""
Tests for table part import/export functionality.

This module tests the import and export services for table parts,
including file format detection, data validation, and error handling.
"""

import unittest
import tempfile
import os
import csv
from datetime import date
from typing import List, Dict, Any

import openpyxl

# Add src to path for imports
import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from services.table_part_import_service import (
    TablePartImportService, ImportColumn, ImportFormat, ImportValidationError
)
from services.table_part_export_service import (
    TablePartExportService, ExportColumn, ExportFormat, ExportOptions
)


class TestTablePartImportService(unittest.TestCase):
    """Test cases for table part import service"""
    
    def setUp(self):
        """Set up test fixtures"""
        self.service = TablePartImportService()
        self.target_columns = [
            ImportColumn(
                source_name="код",
                target_field="code",
                data_type="str",
                required=True
            ),
            ImportColumn(
                source_name="наименование",
                target_field="name",
                data_type="str",
                required=True
            ),
            ImportColumn(
                source_name="количество",
                target_field="quantity",
                data_type="float",
                required=True
            ),
            ImportColumn(
                source_name="цена",
                target_field="price",
                data_type="float",
                required=False,
                default_value=0.0
            )
        ]
    
    def test_detect_file_format(self):
        """Test file format detection"""
        # Test Excel formats
        self.assertEqual(self.service.detect_file_format("test.xlsx"), ImportFormat.EXCEL)
        self.assertEqual(self.service.detect_file_format("test.xls"), ImportFormat.EXCEL)
        
        # Test CSV format
        self.assertEqual(self.service.detect_file_format("test.csv"), ImportFormat.CSV)
        
        # Test unsupported format
        with self.assertRaises(ImportValidationError):
            self.service.detect_file_format("test.txt")
    
    def test_excel_import(self):
        """Test Excel file import"""
        # Create test Excel file
        tmp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        tmp_file.close()  # Close file handle so Excel can write to it
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Write headers
            ws['A1'] = 'код'
            ws['B1'] = 'наименование'
            ws['C1'] = 'количество'
            ws['D1'] = 'цена'
            
            # Write test data
            ws['A2'] = '001'
            ws['B2'] = 'Работа 1'
            ws['C2'] = 10.5
            ws['D2'] = 1500.0
            
            ws['A3'] = '002'
            ws['B3'] = 'Работа 2'
            ws['C3'] = 25.0
            ws['D3'] = 800.0
            
            wb.save(tmp_file.name)
            wb.close()  # Close workbook
            
            # Test preview creation
            preview = self.service.create_preview(tmp_file.name, self.target_columns)
            
            self.assertEqual(preview.detected_format, ImportFormat.EXCEL)
            self.assertEqual(len(preview.headers), 4)
            self.assertEqual(preview.total_rows, 2)
            self.assertIn('код', preview.suggested_mappings)
            
            # Test data import
            column_mappings = {
                'код': 'code',
                'наименование': 'name',
                'количество': 'quantity',
                'цена': 'price'
            }
            
            result = self.service.import_data(
                tmp_file.name,
                column_mappings,
                self.target_columns,
                skip_header=True
            )
            
            self.assertTrue(result.success)
            self.assertEqual(result.imported_rows, 2)
            self.assertEqual(len(result.data), 2)
            
            # Check first row
            first_row = result.data[0]
            self.assertEqual(first_row['code'], '001')
            self.assertEqual(first_row['name'], 'Работа 1')
            self.assertEqual(first_row['quantity'], 10.5)
            self.assertEqual(first_row['price'], 1500.0)
            
        finally:
            try:
                os.unlink(tmp_file.name)
            except (OSError, PermissionError):
                pass  # Ignore file deletion errors
    
    def test_csv_import(self):
        """Test CSV file import"""
        # Create test CSV file
        tmp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8')
        
        try:
            writer = csv.writer(tmp_file)
            
            # Write headers
            writer.writerow(['код', 'наименование', 'количество', 'цена'])
            
            # Write test data
            writer.writerow(['001', 'Работа 1', '10.5', '1500.0'])
            writer.writerow(['002', 'Работа 2', '25.0', '800.0'])
            
            tmp_file.flush()
            tmp_file.close()  # Close file handle
            
            # Test preview creation
            preview = self.service.create_preview(tmp_file.name, self.target_columns)
            
            self.assertEqual(preview.detected_format, ImportFormat.CSV)
            self.assertEqual(len(preview.headers), 4)
            # CSV reader may count empty lines, so just check we have at least 2 rows
            self.assertGreaterEqual(preview.total_rows, 2)
            
            # Test data import
            column_mappings = {
                'код': 'code',
                'наименование': 'name',
                'количество': 'quantity',
                'цена': 'price'
            }
            
            result = self.service.import_data(
                tmp_file.name,
                column_mappings,
                self.target_columns,
                skip_header=True
            )
            
            self.assertTrue(result.success)
            self.assertEqual(result.imported_rows, 2)
            self.assertEqual(len(result.data), 2)
            
        finally:
            try:
                os.unlink(tmp_file.name)
            except (OSError, PermissionError):
                pass  # Ignore file deletion errors
    
    def test_validation_errors(self):
        """Test validation error handling"""
        # Create CSV with missing required field
        tmp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8')
        
        try:
            writer = csv.writer(tmp_file)
            
            # Write headers (missing required 'наименование')
            writer.writerow(['код', 'количество', 'цена'])
            
            # Write test data
            writer.writerow(['001', '10.5', '1500.0'])
            
            tmp_file.flush()
            tmp_file.close()  # Close file handle
            
            column_mappings = {
                'код': 'code',
                'количество': 'quantity',
                'цена': 'price'
            }
            
            result = self.service.import_data(
                tmp_file.name,
                column_mappings,
                self.target_columns,
                skip_header=True
            )
            
            # Should have errors due to missing required field
            self.assertEqual(result.failed_rows, 1)
            self.assertTrue(len(result.errors) > 0)
            
        finally:
            try:
                os.unlink(tmp_file.name)
            except (OSError, PermissionError):
                pass  # Ignore file deletion errors


class TestTablePartExportService(unittest.TestCase):
    """Test cases for table part export service"""
    
    def setUp(self):
        """Set up test fixtures"""
        self.service = TablePartExportService()
        self.test_data = [
            {
                'code': '001',
                'name': 'Работа 1',
                'quantity': 10.5,
                'price': 1500.0,
                'sum': 15750.0,
                'date': date(2024, 1, 15)
            },
            {
                'code': '002',
                'name': 'Работа 2',
                'quantity': 25.0,
                'price': 800.0,
                'sum': 20000.0,
                'date': date(2024, 1, 16)
            }
        ]
        
        self.export_columns = [
            ExportColumn(
                field_name="code",
                display_name="Код",
                data_type="str"
            ),
            ExportColumn(
                field_name="name",
                display_name="Наименование",
                data_type="str"
            ),
            ExportColumn(
                field_name="quantity",
                display_name="Количество",
                data_type="float"
            ),
            ExportColumn(
                field_name="price",
                display_name="Цена",
                data_type="float"
            ),
            ExportColumn(
                field_name="sum",
                display_name="Сумма",
                data_type="float"
            ),
            ExportColumn(
                field_name="date",
                display_name="Дата",
                data_type="date"
            )
        ]
    
    def test_excel_export(self):
        """Test Excel file export"""
        tmp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        tmp_file.close()  # Close file handle
        
        try:
            options = ExportOptions(
                include_headers=True,
                apply_formatting=True,
                auto_fit_columns=True
            )
            
            result = self.service.export_data(
                self.test_data,
                tmp_file.name,
                self.export_columns,
                options
            )
            
            self.assertTrue(result.success)
            self.assertEqual(result.exported_rows, 2)
            self.assertEqual(result.file_path, tmp_file.name)
            
            # Verify file was created and has content
            self.assertTrue(os.path.exists(tmp_file.name))
            
            # Read back and verify content
            wb = openpyxl.load_workbook(tmp_file.name)
            ws = wb.active
            
            # Check headers
            self.assertEqual(ws['A1'].value, 'Код')
            self.assertEqual(ws['B1'].value, 'Наименование')
            
            # Check first data row
            self.assertEqual(ws['A2'].value, '001')
            self.assertEqual(ws['B2'].value, 'Работа 1')
            self.assertEqual(ws['C2'].value, 10.5)
            
            wb.close()  # Close workbook
            
        finally:
            try:
                os.unlink(tmp_file.name)
            except (OSError, PermissionError):
                pass  # Ignore file deletion errors
    
    def test_csv_export(self):
        """Test CSV file export"""
        tmp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False)
        tmp_file.close()  # Close file handle
        
        try:
            options = ExportOptions(
                include_headers=True,
                encoding='utf-8',
                delimiter=','
            )
            
            result = self.service.export_data(
                self.test_data,
                tmp_file.name,
                self.export_columns,
                options
            )
            
            self.assertTrue(result.success)
            self.assertEqual(result.exported_rows, 2)
            
            # Verify file was created and has content
            self.assertTrue(os.path.exists(tmp_file.name))
            
            # Read back and verify content
            with open(tmp_file.name, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                rows = list(reader)
            
            # Check headers
            self.assertEqual(rows[0][0], 'Код')
            self.assertEqual(rows[0][1], 'Наименование')
            
            # Check first data row
            self.assertEqual(rows[1][0], '001')
            self.assertEqual(rows[1][1], 'Работа 1')
            
        finally:
            try:
                os.unlink(tmp_file.name)
            except (OSError, PermissionError):
                pass  # Ignore file deletion errors
    
    def test_empty_data_export(self):
        """Test export with empty data"""
        tmp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        tmp_file.close()  # Close file handle
        
        try:
            result = self.service.export_data(
                [],  # Empty data
                tmp_file.name,
                self.export_columns,
                ExportOptions()
            )
            
            self.assertFalse(result.success)
            self.assertTrue(len(result.errors) > 0)
            self.assertIn("Нет данных для экспорта", result.errors[0])
            
        finally:
            try:
                os.unlink(tmp_file.name)
            except (OSError, PermissionError):
                pass  # Ignore file deletion errors
    
    def test_suggested_filename(self):
        """Test suggested filename generation"""
        excel_name = self.service.get_suggested_filename("test_table", ExportFormat.EXCEL)
        csv_name = self.service.get_suggested_filename("test_table", ExportFormat.CSV)
        
        self.assertTrue(excel_name.startswith("test_table_"))
        self.assertTrue(excel_name.endswith(".xlsx"))
        
        self.assertTrue(csv_name.startswith("test_table_"))
        self.assertTrue(csv_name.endswith(".csv"))


def main():
    """Run all tests"""
    # Create test suite
    loader = unittest.TestLoader()
    test_suite = unittest.TestSuite()
    
    # Add import service tests
    test_suite.addTests(loader.loadTestsFromTestCase(TestTablePartImportService))
    
    # Add export service tests
    test_suite.addTests(loader.loadTestsFromTestCase(TestTablePartExportService))
    
    # Run tests
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(test_suite)
    
    # Return exit code based on test results
    return 0 if result.wasSuccessful() else 1


if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)